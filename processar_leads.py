"""
Processar leads do Apify - Gera CSV final
==========================================
Baixa os dados ja coletados (sem gastar creditos),
processa, filtra e gera CSV/Excel com leads qualificados.
"""

import csv
import json
import re
import sys
import os
import requests
from pathlib import Path
from datetime import datetime
from collections import Counter

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

APIFY_TOKEN = "apify_api_g35riRIYbexQigTpDwRm4GJ40emppP36z2dc"
OUTPUT_DIR = Path(__file__).parent / "output" / "apify"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Datasets ja coletados
DATASETS = {
    "dentista": "xMDbuekYSds4w6dXd",
    "advogado": "dF7T0nddVYqR2wO72",
    "clínica estética": "yHgfK9WfVP7o91gHN",
}

CIDADES = [
    "Belford Roxo", "Duque de Caxias", "Guapimirim", "Itaguaí",
    "Japeri", "Magé", "Mesquita", "Nilópolis", "Nova Iguaçu",
    "Paracambi", "Queimados", "São João de Meriti", "Seropédica",
]

FRANQUIAS = [
    "mcdonald", "burguer king", "subway", "habib", "giraffas",
    "bob's", "kfc", "pizza hut", "domino", "grenal",
    "o boticário", "americanas", "casas bahia", "carrefour",
    "assaí", "atacadão", "magazine luiza", "ri happy", "smart fit",
    "odontocompany", "oral un", "dental estética",
]

EMAIL_BLACKLIST = {
    "test@test.com", "email@email.com", "example@example.com",
    "noreply@google.com", "no-reply@google.com",
}

FIELDNAMES = [
    "nome", "nicho", "cidade", "endereco", "telefone", "whatsapp",
    "website", "email", "avaliacao", "num_avaliacoes",
    "instagram", "facebook", "observacao", "relevancia",
]


def validate_email(email):
    if not email:
        return None
    email = str(email).strip().lower()
    if email.startswith("#"):
        # Cloudflare encoded - tentar decodificar
        return None
    if email in EMAIL_BLACKLIST:
        return None
    if "quemfazsite" in email or "placeholder" in email:
        return None
    if len(email) > 80 or len(email) < 6:
        return None
    if not re.match(r"^[\w.+-]+@[\w-]+\.[\w.-]+$", email):
        return None
    ext = "." + email.rsplit(".", 1)[-1] if "." in email else ""
    if ext in {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico"}:
        return None
    return email


def extract_whatsapp(was):
    """Extrai numero do WhatsApp da URL."""
    if not was:
        return ""
    nums = []
    for w in was:
        if isinstance(w, str):
            match = re.search(r"phone=(\d+)", w)
            if match:
                num = match.group(1)
                # Formatar: +55 21 9XXXX-XXXX
                if len(num) >= 12:
                    nums.append(f"+{num[:2]} {num[2:4]} {num[4:8]}-{num[8:]}")
                else:
                    nums.append(f"+{num}")
    return nums[0] if nums else ""


def extract_instagram(instas):
    if not instas:
        return ""
    for i in instas:
        if isinstance(i, str) and "instagram.com/" in i:
            # Pegar o handle
            match = re.search(r"instagram\.com/([a-zA-Z0-9_.]+)", i)
            if match:
                handle = match.group(1)
                if handle not in ["share", "explore", "p", "reel", "stories"]:
                    return f"@{handle}"
    return ""


def extract_facebook(fbs):
    if not fbs:
        return ""
    for f in fbs:
        if isinstance(f, str) and "facebook.com/" in f:
            match = re.search(r"facebook\.com/([a-zA-Z0-9_.]+)", f)
            if match:
                handle = match.group(1)
                if handle not in ["share.php", "sharer", "share"]:
                    return f
    return ""


def detect_cidade(item):
    """Detecta a cidade do endereço ou campo city."""
    # Campo city direto
    city = (item.get("city") or "").lower()
    for c in CIDADES:
        if c.lower() in city:
            return c

    # Campo address
    addr = (item.get("address") or "").lower()
    for c in CIDADES:
        if c.lower() in addr:
            return c

    return ""


def normalize(item, nicho):
    """Normaliza um lead."""
    nome = item.get("title") or ""
    if not nome:
        return None

    # Detectar cidade
    cidade = detect_cidade(item)
    if not cidade:
        return None  # Fora da Baixada

    endereco = item.get("address") or ""
    telefone = item.get("phone") or ""
    website = item.get("website") or ""
    avaliacao = item.get("totalScore") or ""
    num_avaliacoes = item.get("reviewsCount") or ""
    domain = item.get("domain") or ""

    # Emails
    emails_raw = item.get("emails") or []
    email = ""
    for e in emails_raw:
        valid = validate_email(e)
        if valid:
            email = valid
            break

    # WhatsApp
    whatsapp = extract_whatsapp(item.get("whatsapps") or [])

    # Redes sociais
    instagram = extract_instagram(item.get("instagrams") or [])
    facebook = extract_facebook(item.get("facebooks") or [])

    # Franquia
    nome_lower = nome.lower()
    is_franchise = any(f in nome_lower for f in FRANQUIAS)

    # Site profissional (heurística: tem domínio próprio)
    has_pro_site = bool(domain and "." in domain and
                       not any(s in domain for s in ["facebook", "instagram", "wix"]))

    # Fechado permanentemente
    if item.get("permanentlyClosed"):
        return None

    # Observações
    obs = []
    if not website:
        obs.append("SEM SITE - oportunidade!")
    elif "facebook" in (website or "").lower():
        obs.append("Site é só Facebook")
    elif "instagram" in (website or "").lower():
        obs.append("Site é só Instagram")
    if not email:
        obs.append("sem email")
    if not telefone and not whatsapp:
        obs.append("sem telefone")
    if not instagram:
        obs.append("sem Instagram")
    if whatsapp and not website:
        obs.append("só WhatsApp, sem site")

    # Score de relevância
    score = 0
    if telefone:
        score += 2
    if whatsapp:
        score += 2
    if email:
        score += 3
    if website and not has_pro_site:
        score += 1
    elif not website:
        score += 2  # Sem site = melhor prospecto
    if instagram:
        score += 1
    try:
        if avaliacao and float(avaliacao) >= 4.0:
            score += 1
    except (ValueError, TypeError):
        pass
    try:
        if num_avaliacoes and int(num_avaliacoes) >= 10:
            score += 1
    except (ValueError, TypeError):
        pass

    return {
        "nome": nome,
        "nicho": nicho,
        "cidade": cidade,
        "endereco": endereco,
        "telefone": telefone,
        "whatsapp": whatsapp,
        "website": website,
        "email": email,
        "avaliacao": avaliacao,
        "num_avaliacoes": num_avaliacoes,
        "instagram": instagram,
        "facebook": facebook,
        "observacao": " | ".join(obs),
        "relevancia": score,
        "_is_franchise": is_franchise,
        "_has_pro_site": has_pro_site,
    }


def filter_leads(leads, target=100):
    """Filtra leads pelas regras do usuário."""
    # Sem nome
    filtered = [l for l in leads if l]

    # Sem franquias
    filtered = [l for l in filtered if not l.get("_is_franchise")]

    # Remover duplicatas (mesmo nome + cidade)
    seen = set()
    unique = []
    for l in filtered:
        key = f"{l['nome'].lower().strip()}|{l['cidade'].lower().strip()}"
        if key not in seen:
            seen.add(key)
            unique.append(l)

    # Ordenar por relevância
    unique.sort(key=lambda x: x.get("relevancia", 0), reverse=True)

    # Top N
    result = unique[:target]

    # Limpar campos internos
    for l in result:
        l.pop("_is_franchise", None)
        l.pop("_has_pro_site", None)

    return result, len(unique)


def export_csv(leads, filename):
    filepath = OUTPUT_DIR / filename
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(leads)
    return filepath


def export_excel(leads, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Qualificados"

        headers_map = {
            "nome": "Nome",
            "nicho": "Nicho",
            "cidade": "Cidade",
            "endereco": "Endereço",
            "telefone": "Telefone",
            "whatsapp": "WhatsApp",
            "website": "Website",
            "email": "Email",
            "avaliacao": "Avaliação",
            "num_avaliacoes": "Nº Avaliações",
            "instagram": "Instagram",
            "facebook": "Facebook",
            "observacao": "Observação (argumento de venda)",
            "relevancia": "Score",
        }

        headers = list(headers_map.values())
        fields = list(headers_map.keys())

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        yellow_fill = PatternFill("solid", fgColor="FEF08A")
        green_fill = PatternFill("solid", fgColor="D1FAE5")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, lead in enumerate(leads, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=lead.get(field, ""))

            # Sem site = amarelo (melhor prospecto)
            if not lead.get("website"):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = yellow_fill
            # Com email = verde na coluna email
            if lead.get("email"):
                ws.cell(row=row_idx, column=8).fill = green_fill

        widths = [35, 18, 22, 45, 18, 20, 35, 30, 10, 12, 25, 35, 40, 8]
        for i, w in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        filepath = OUTPUT_DIR / filename
        wb.save(filepath)
        return filepath
    except ImportError:
        print("  [!] openpyxl nao instalado - CSV apenas")
        return None


def main():
    print("=" * 60)
    print("PROCESSANDO LEADS - DADOS JA COLETADOS (sem custo)")
    print("=" * 60)

    all_leads = []

    for nicho, dataset_id in DATASETS.items():
        print(f"\nBaixando {nicho.upper()} (dataset: {dataset_id})...")
        url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={APIFY_TOKEN}&clean=true&format=json"
        resp = requests.get(url, timeout=60)
        items = resp.json()
        print(f"  Resultados: {len(items)}")

        for item in items:
            lead = normalize(item, nicho)
            if lead:
                all_leads.append(lead)

    print(f"\nTotal bruto na Baixada: {len(all_leads)} leads")

    # Filtrar
    filtered, total_unique = filter_leads(all_leads, target=100)
    print(f"Apos filtros: {len(filtered)} leads (de {total_unique} unicos)")

    # Exportar
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = export_csv(filtered, f"leads_baixada_{timestamp}.csv")
    print(f"\nCSV: {csv_file}")

    xlsx_file = export_excel(filtered, f"leads_baixada_{timestamp}.xlsx")
    if xlsx_file:
        print(f"Excel: {xlsx_file}")

    # Resumo
    print(f"\n{'='*60}")
    print("RESUMO")
    print(f"{'='*60}")
    print(f"Leads qualificados: {len(filtered)}")

    print(f"\nPor nicho:")
    for n, c in Counter(l["nicho"] for l in filtered).most_common():
        print(f"  {n}: {c}")

    print(f"\nPor cidade:")
    for c, n in Counter(l["cidade"] for l in filtered).most_common():
        print(f"  {c}: {n}")

    com_tel = sum(1 for l in filtered if l["telefone"])
    com_wa = sum(1 for l in filtered if l["whatsapp"])
    com_email = sum(1 for l in filtered if l["email"])
    com_insta = sum(1 for l in filtered if l["instagram"])
    sem_site = sum(1 for l in filtered if not l["website"])

    print(f"\nCom telefone: {com_tel}/{len(filtered)}")
    print(f"Com WhatsApp: {com_wa}/{len(filtered)}")
    print(f"Com email: {com_email}/{len(filtered)}")
    print(f"Com Instagram: {com_insta}/{len(filtered)}")
    print(f"Sem site (melhores prospectos): {sem_site}/{len(filtered)}")

    # Mostrar top 10
    print(f"\nTOP 10 LEADS:")
    for i, l in enumerate(filtered[:10], 1):
        print(f"  {i}. {l['nome']} ({l['nicho']}) - {l['cidade']}")
        print(f"     Tel: {l['telefone']} | WA: {l['whatsapp']} | Email: {l['email']}")
        print(f"     Obs: {l['observacao']}")


if __name__ == "__main__":
    main()
