"""
Capturador de Emails - Baixada Fluminense, RJ
==============================================
Busca DOIS emails por comercio:
  1. Email CNPJ (pessoal do dono, via sites de consulta)
  2. Email comercial (onde o comercio atende, via web)
Pode ser interrompido (Ctrl+C) e retoma de onde parou.

Uso: python capturar_emails.py
"""

import asyncio
import csv
import json
import os
import random
import re
import sys
import urllib.parse
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent / "output"
CSV_FILE = OUTPUT_DIR / "todos_comercios.csv"
PROGRESS_FILE = OUTPUT_DIR / "progresso_emails.json"

# Config
DELAY_MIN = 3.0
DELAY_MAX = 6.0
HEADLESS = False
BATCH_SAVE = 10
MAX_RETRIES = 3

# Sites de consulta CNPJ
CNPJ_SITES = [
    "cnpj.biz",
    "consultacnpj.com",
    "cnpj.info",
    "cnpj.rocks",
    "do.cnpj.biz",
]


# ── Progress / Resume ──────────────────────────────────────────────

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"indice_atual": 0, "total": 0, "emails_cnpj": 0, "emails_comercial": 0, "resultados": []}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ── Email Extraction ───────────────────────────────────────────────

BLACKLIST = {
    "test@test.com", "email@email.com", "example@example.com",
    "contato@exemplo.com", "teste@teste.com", "user@domain.com",
    "noreply@google.com", "no-reply@google.com",
    "newsletter@googlegroups.com", "mailer-daemon@google.com",
}

IMG_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico"}


def extract_emails(text):
    """Extrai emails validos de um texto. Aceita qualquer provedor."""
    if not text:
        return []

    t = text.lower()
    t = t.replace("[at]", "@").replace("(at)", "@").replace("{at}", "@")
    t = t.replace("[arroba]", "@").replace("(arroba)", "@")
    t = t.replace(" at ", "@")
    t = t.replace("[dot]", ".").replace("(dot)", ".")
    t = t.replace(" dot ", ".")

    raw = re.findall(r"[\w.+-]+@[\w-]+\.[\w.-]+", t)

    valid = []
    seen = set()
    for email in raw:
        email = email.strip(".")

        if len(email) > 80 or email in BLACKLIST:
            continue

        domain = email.split("@")[-1] if "@" in email else ""
        if "." in domain:
            ext = "." + domain.rsplit(".", 1)[-1]
            if ext in IMG_EXTENSIONS:
                continue

        if len(domain) < 4:
            continue

        if email not in seen:
            seen.add(email)
            valid.append(email)

    return valid


# ── Etapa 1: Email CNPJ (pessoal do dono) ──────────────────────────

async def buscar_email_cnpj(page, nome, cidade, blacklist_emails):
    """Busca email CNPJ direto no cnpj.biz. Retorna email ou vazio."""
    cidade_clean = cidade.replace(", RJ", "").replace(",RJ", "").strip()

    try:
        # Vai direto na busca de empresas do cnpj.biz
        await page.goto("https://cnpj.biz/empresas", wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(2)

        # Verifica se a pagina carregou (nao foi bloqueada)
        body_check = await page.inner_text("body")
        if any(p in body_check.lower() for p in ["navegador", "browser is too old", "cloudflare"]):
            return ""

        # Preenche o campo de busca
        search_input = page.locator('input[type="text"], input[type="search"], input[name="q"], '
                                     'input[placeholder*="Buscar"], input[placeholder*="empresa"], '
                                     'input[placeholder*="nome"]').first
        if await search_input.count() == 0:
            return ""

        await search_input.click()
        await asyncio.sleep(0.3)
        await search_input.fill(nome)
        await asyncio.sleep(0.5)
        await search_input.press("Enter")
        await asyncio.sleep(3)

        # Pega todo o HTML da pagina pra analisar
        body = await page.inner_text("body")

        # Se nao tem ATIVA, nao achou resultados
        if "ATIVA" not in body.upper():
            return ""

        # Estrategia simples: pega TODOS os links da pagina
        # Encontra o link cujo texto contem parte do nome da empresa
        # e que esteja proximo de "ATIVA" e da cidade
        todos_links = page.locator("a")
        total_links = await todos_links.count()

        nome_lower = nome.lower()
        tentativas_max = 10  # Limita tentativas pra evitar loop infinito
        tentativas = 0

        for idx in range(total_links):
            if tentativas >= tentativas_max:
                print(f"    [!] Limite de tentativas atingido para {nome[:30]}")
                break

            try:
                link = todos_links.nth(idx)
                link_text = (await link.inner_text()).strip()

                # Verifica se o link tem a ver com o nome do comercio
                if len(link_text) < 5:
                    continue

                # Pega o HTML do parente proximo pra ver contexto
                parent = link.locator("xpath=ancestor::*[position() <= 3]")
                parent_text = ""
                try:
                    parent_text = await parent.first.inner_text()
                except Exception:
                    parent_text = link_text

                parent_upper = parent_text.upper()

                # Pula se BAIXADA
                if "BAIXADA" in parent_upper:
                    continue

                # Aceita se tem ATIVA + cidade
                if "ATIVA" in parent_upper and cidade_clean.lower() in parent_text.lower():
                    tentativas += 1
                    print(f"    [{tentativas}/{tentativas_max}] Tentando: {link_text[:40]}")

                    # Clica!
                    await link.click()
                    await asyncio.sleep(3)

                    # Extrai email da pagina de detalhes
                    email = await extrair_email_detalhes(page, blacklist_emails)
                    if email:
                        return email

                    # Volta pra lista
                    await page.go_back(timeout=10000)
                    await asyncio.sleep(2)
            except Exception as e:
                print(f"    [!] Erro ao processar link: {str(e)[:40]}")
                try:
                    await page.go_back(timeout=5000)
                    await asyncio.sleep(1)
                except Exception:
                    pass

        return ""

    except Exception as e:
        print(f"    [!] Erro em buscar_email_cnpj: {str(e)[:50]}")
        return ""


async def extrair_email_detalhes(page, blacklist_emails):
    """Na pagina de detalhes, intercepta a API do cnpj.biz para pegar o email."""
    try:
        await page.wait_for_load_state("networkidle", timeout=10000)
        await asyncio.sleep(2)

        # FECHAR qualquer popup/modal/overlay
        await page.keyboard.press("Escape")
        await asyncio.sleep(0.5)
        await page.evaluate("""() => {
            // Remove todos os overlays e popups
            document.querySelectorAll(
                '[class*="modal"], [class*="popup"], [class*="overlay"], [class*="notification"], [class*="banner"], [class*="subscribe"]'
            ).forEach(el => el.remove());
        }""")
        await asyncio.sleep(0.5)

        # PRIMEIRO: Verifica se o email ja esta no HTML completo
        html_completo = await page.content()
        emails = extract_emails(html_completo)
        emails = [e for e in emails if e not in blacklist_emails]
        if emails:
            print(f"    [✓] Email no HTML: {emails[0]}")
            return emails[0]

        # SEGUNDO: Intercepta requisicoes de rede para pegar email da API
        print(f"    [  ] Interceptando rede...")
        email_capturado = []

        async def capturar_email_response(response):
            try:
                url = response.url
                if "email" in url.lower() or "contact" in url.lower() or "empresa" in url.lower():
                    body = await response.text()
                    encontrados = extract_emails(body)
                    encontrados = [e for e in encontrados if e not in blacklist_emails]
                    if encontrados:
                        email_capturado.append(encontrados[0])
            except Exception:
                pass

        page.on("response", capturar_email_response)

        # TERCEIRO: Scroll ate a secao de contatos e clica no Ver E-mail
        print(f"    [  ] Procurando botao Ver E-mail...")

        # Scroll ate encontrar o texto "E-mail" na pagina
        await page.evaluate("""() => {
            // Procura o texto "E-mail" e faz scroll ate ele
            const walker = document.createTreeWalker(
                document.body,
                NodeFilter.SHOW_TEXT,
                {
                    acceptNode: function(node) {
                        if (node.textContent.includes('E-mail') || node.textContent.includes('Email')) {
                            return NodeFilter.FILTER_ACCEPT;
                        }
                        return NodeFilter.FILTER_SKIP;
                    }
                }
            );
            const node = walker.nextNode();
            if (node) {
                node.parentElement.scrollIntoView({behavior: 'smooth', block: 'center'});
            }
        }""")
        await asyncio.sleep(2)

        # Fecha popup novamente apos scroll
        await page.keyboard.press("Escape")
        await asyncio.sleep(0.3)

        # Agora clica no link "Ver E-mail" que esta ao lado do email mascarado
        clicou = await page.evaluate("""() => {
            // Procura especificamente pelo texto "(Ver E-mail)"
            const allElements = document.querySelectorAll('a, span, button, div, p');
            for (let el of allElements) {
                const text = el.textContent || '';
                // Procura o texto exato "Ver E-mail" ou "(Ver E-mail)"
                if (text.includes('Ver E-mail') || text.includes('Ver Email') || text.includes('ver e-mail')) {
                    // Remove qualquer overlay que possa estar em cima
                    document.querySelectorAll('[style*="z-index"]').forEach(o => {
                        const s = window.getComputedStyle(o);
                        if (s.position === 'fixed' && parseInt(s.zIndex) > 100) {
                            o.remove();
                        }
                    });

                    el.scrollIntoView({behavior: 'smooth', block: 'center'});
                    el.click();
                    return true;
                }
            }
            return false;
        }""")

        if clicou:
            print(f"    [✓] Clicou no Ver E-mail!")
            await asyncio.sleep(4)

            # Verifica se interceptou o email via rede
            if email_capturado:
                print(f"    [✓] Email via rede: {email_capturado[0]}")
                page.remove_listener("response", capturar_email_response)
                return email_capturado[0]

            # Verifica se o email apareceu no HTML
            html_novo = await page.content()
            emails_novos = extract_emails(html_novo)
            emails_novos = [e for e in emails_novos if e not in blacklist_emails]
            if emails_novos and emails_novos != emails:
                print(f"    [✓] Email revelado: {emails_novos[0]}")
                page.remove_listener("response", capturar_email_response)
                return emails_novos[0]

        page.remove_listener("response", capturar_email_response)

        print(f"    [-] Nenhum email encontrado")
        return ""

    except Exception as e:
        print(f"    [!] Erro: {str(e)[:50]}")
        return ""


# ── Etapa 2: Email comercial (onde o comercio atende) ──────────────

async def buscar_email_comercial(page, nome, cidade, blacklist_emails, email_cnpj=""):
    """Busca email comercial do comercio via Bing e diretamente em sites."""
    cidade_clean = cidade.replace(", RJ", "").replace(",RJ", "").strip()
    # Adiciona o email CNPJ na blacklist pra nao repetir
    bl = set(blacklist_emails)
    if email_cnpj:
        bl.add(email_cnpj)

    print(f"    [→] Buscando email comercial para: {nome[:40]}")

    # ESTRATEGIA 1: Busca no Bing com varias queries
    queries = [
        f'"{nome}" {cidade_clean} email',
        f'"{nome}" {cidade_clean} contato',
        f'"{nome}" {cidade_clean} @',
        f"{nome} {cidade_clean} site oficial",
    ]

    for query_idx, query in enumerate(queries):
        try:
            url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}"
            print(f"    [  ] Busca {query_idx + 1}: {query[:50]}")

            await page.goto(url, wait_until="domcontentloaded", timeout=20000)
            await asyncio.sleep(1.5)

            # Extrai emails dos snippets e titulos
            textos = []

            # Seletores do Bing para snippets
            seletores_bing = [
                ".b_caption p", ".b_lineclamp2", "li.b_algo p",
                ".b_snippet", "p.snippet", ".algoSnippet",
            ]

            for sel in seletores_bing:
                try:
                    snippets = page.locator(sel)
                    count = await snippets.count()
                    for j in range(min(count, 20)):
                        try:
                            txt = await snippets.nth(j).inner_text()
                            textos.append(txt)
                        except Exception:
                            pass
                except Exception:
                    pass

            # Titulos
            try:
                titulos = page.locator("li.b_algo h2, h2 a, .b_title")
                count_t = await titulos.count()
                for j in range(min(count_t, 15)):
                    try:
                        txt = await titulos.nth(j).inner_text()
                        textos.append(txt)
                    except Exception:
                        pass
            except Exception:
                pass

            # Body completo
            try:
                body = await page.inner_text("body")
                textos.append(body)
            except Exception:
                pass

            # Extrai emails
            texto_completo = " ".join(textos)
            emails = extract_emails(texto_completo)
            emails = [e for e in emails if e not in bl]

            if emails:
                print(f"    [✓] Email encontrado no Bing: {emails[0]}")
                return emails[0]

            # ESTRATEGIA 2: Clica no primeiro resultado relevante e extrai email do site
            try:
                # Pega o primeiro link de resultado
                primeiro_link = page.locator("li.b_algo h2 a").first
                if await primeiro_link.count() > 0:
                    href = await primeiro_link.get_attribute("href")
                    if href and not any(x in href for x in ["linkedin", "facebook", "instagram", "twitter"]):
                        print(f"    [  ] Visitando site: {href[:60]}")
                        await page.goto(href, wait_until="domcontentloaded", timeout=15000)
                        await asyncio.sleep(2)

                        # Procura email no site
                        body_site = await page.inner_text("body")
                        emails_site = extract_emails(body_site)
                        emails_site = [e for e in emails_site if e not in bl]

                        if emails_site:
                            print(f"    [✓] Email encontrado no site: {emails_site[0]}")
                            return emails_site[0]

                        # Volta pro Bing
                        await page.go_back(timeout=10000)
                        await asyncio.sleep(1)
            except Exception:
                pass

        except Exception as e:
            print(f"    [!] Erro na busca {query_idx + 1}: {str(e)[:40]}")
            await asyncio.sleep(2)

    print(f"    [-] Nenhum email comercial encontrado")
    return ""


# ── CSV Export ──────────────────────────────────────────────────────

def update_csvs(resultados):
    """Atualiza os CSVs com os dois emails encontrados."""
    email_map = {}
    for r in resultados:
        key = f"{r['nome'].lower().strip()}|{r.get('cidade', '').lower().strip()}"
        if key not in email_map:
            email_map[key] = {}
        if r.get("email_cnpj"):
            email_map[key]["email_cnpj"] = r["email_cnpj"]
        if r.get("email_comercial"):
            email_map[key]["email_comercial"] = r["email_comercial"]

    fieldnames = [
        "nome", "endereco", "telefone", "email_cnpj", "email_comercial", "categoria",
        "tem_site", "url_site", "avaliacao", "num_avaliacoes", "cidade",
    ]
    todos = []
    with open(CSV_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = f"{row['nome'].lower().strip()}|{row.get('cidade', '').lower().strip()}"
            if key in email_map:
                if email_map[key].get("email_cnpj") and not row.get("email_cnpj"):
                    row["email_cnpj"] = email_map[key]["email_cnpj"]
                if email_map[key].get("email_comercial") and not row.get("email_comercial"):
                    row["email_comercial"] = email_map[key]["email_comercial"]
            # Garante que as colunas existem
            if "email_cnpj" not in row:
                row["email_cnpj"] = ""
            if "email_comercial" not in row:
                row["email_comercial"] = ""
            todos.append(row)

    with open(CSV_FILE, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(todos)
    print(f"  [atualizado] todos_comercios.csv")

    sem_site = [c for c in todos if c.get("tem_site") == "False"]
    with open(OUTPUT_DIR / "leads_sem_site.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(sem_site)
    print(f"  [atualizado] leads_sem_site.csv")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads sem Site"

        headers = [
            "Nome", "Endereco", "Telefone", "Email CNPJ", "Email Comercial",
            "Categoria", "Tem Site?", "URL do Site", "Avaliacao", "N Avaliacoes", "Cidade",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, b in enumerate(sem_site, 2):
            ws.cell(row=row_idx, column=1, value=b["nome"])
            ws.cell(row=row_idx, column=2, value=b["endereco"])
            ws.cell(row=row_idx, column=3, value=b["telefone"])
            ws.cell(row=row_idx, column=4, value=b.get("email_cnpj", ""))
            ws.cell(row=row_idx, column=5, value=b.get("email_comercial", ""))
            ws.cell(row=row_idx, column=6, value=b["categoria"])
            ws.cell(row=row_idx, column=7, value="Nao")
            ws.cell(row=row_idx, column=8, value=b["url_site"])
            ws.cell(row=row_idx, column=9, value=b["avaliacao"])
            ws.cell(row=row_idx, column=10, value=b["num_avaliacoes"])
            ws.cell(row=row_idx, column=11, value=b.get("cidade", ""))

        widths = [35, 45, 18, 30, 30, 25, 10, 35, 10, 12, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        wb.save(OUTPUT_DIR / "leads_sem_site.xlsx")
        print(f"  [atualizado] leads_sem_site.xlsx")
    except ImportError:
        pass


# ── Main ────────────────────────────────────────────────────────────

async def main():
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("[!] Playwright nao encontrado. Instalando...")
        os.system(f"{sys.executable} -m pip install playwright")
        os.system(f"{sys.executable} -m playwright install chromium")
        from playwright.async_api import async_playwright

    if not CSV_FILE.exists():
        print(f"[!] Arquivo nao encontrado: {CSV_FILE}")
        return

    comercios = []
    with open(CSV_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            comercios.append(row)

    total = len(comercios)
    print("=" * 60)
    print("  CAPTURADOR DE EMAILS - BAIXADA FLUMINENSE, RJ")
    print(f"  {total} comercios para buscar")
    print("  ETAPA 1: Email CNPJ (pessoal do dono)")
    print("  ETAPA 2: Email comercial (onde o comercio atende)")
    print("=" * 60)

    progress = load_progress()
    start = progress["indice_atual"]

    # Pula comercios ja processados (tem pelo menos um email)
    ja_processados = set()
    for r in progress["resultados"]:
        if r.get("email_cnpj") or r.get("email_comercial"):
            key = f"{r['nome'].lower().strip()}|{r.get('cidade', '').lower().strip()}"
            ja_processados.add(key)

    print(f"  Ja processados: {len(ja_processados)}")
    print(f"  Emails CNPJ anteriores: {progress.get('emails_cnpj', 0)}")
    print(f"  Emails comerciais anteriores: {progress.get('emails_comercial', 0)}")
    print(f"  Retomando do indice: {start}")
    print("=" * 60)

    blacklist_emails = set(BLACKLIST)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1366, "height": 768},
            locale="pt-BR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            ),
        )

        # Bloqueia notificacoes push do navegador
        await ctx.route("**/sw.js", lambda route: route.abort())
        await ctx.route("**/manifest.json", lambda route: route.abort())
        await ctx.route("**/firebase-messaging-sw.js", lambda route: route.abort())
        await ctx.route("**/OneSignalSDK*", lambda route: route.abort())
        await ctx.route("**/push*", lambda route: route.abort())

        page = await ctx.new_page()

        # Handler para fechar alerts nativos do navegador
        page.on("dialog", lambda dialog: dialog.dismiss())

        try:
            await page.goto("https://www.bing.com", wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(2)
        except Exception:
            pass

        try:
            for i in range(start, total):
                comercio = comercios[i]
                nome = comercio["nome"]
                cidade = comercio.get("cidade", "")

                key = f"{nome.lower().strip()}|{cidade.lower().strip()}"
                if key in ja_processados:
                    progress["indice_atual"] = i + 1
                    continue

                # ETAPA 1: Email CNPJ (pessoal do dono)
                email_cnpj = await buscar_email_cnpj(page, nome, cidade, blacklist_emails)

                # ETAPA 2: Email comercial (onde o comercio atende)
                email_comercial = await buscar_email_comercial(page, nome, cidade, blacklist_emails, email_cnpj)

                # Log
                partes = []
                if email_cnpj:
                    progress["emails_cnpj"] = progress.get("emails_cnpj", 0) + 1
                    partes.append(f"CNPJ: {email_cnpj}")
                if email_comercial:
                    progress["emails_comercial"] = progress.get("emails_comercial", 0) + 1
                    partes.append(f"COM: {email_comercial}")

                if partes:
                    ja_processados.add(key)
                    print(f"  [{i + 1}/{total}] {' | '.join(partes)} - {nome[:40]}")
                else:
                    print(f"  [{i + 1}/{total}] sem email - {nome[:40]}")

                progress["resultados"].append({
                    "nome": nome,
                    "cidade": cidade,
                    "email_cnpj": email_cnpj,
                    "email_comercial": email_comercial,
                })

                if (i + 1) % BATCH_SAVE == 0:
                    progress["indice_atual"] = i + 1
                    progress["total"] = total
                    save_progress(progress)
                    total_emails = progress.get("emails_cnpj", 0) + progress.get("emails_comercial", 0)
                    print(f"  [progresso salvo] {i + 1}/{total} | {total_emails} emails")

                delay = random.uniform(DELAY_MIN, DELAY_MAX)
                await asyncio.sleep(delay)

        except KeyboardInterrupt:
            print(f"\n\n  [INTERROMPIDO] Progresso salvo no indice {i}")
            print(f"  Para continuar: python capturar_emails.py")

        finally:
            progress["indice_atual"] = progress.get("indice_atual", total)
            progress["total"] = total
            save_progress(progress)
            await browser.close()

    print("\nAtualizando CSVs...")
    update_csvs(progress["resultados"])

    total_emails = progress.get("emails_cnpj", 0) + progress.get("emails_comercial", 0)
    print(f"\n{'=' * 60}")
    print(f"  RESUMO")
    print(f"  Processados:         {progress['indice_atual']}/{total}")
    print(f"  Emails CNPJ:         {progress.get('emails_cnpj', 0)}")
    print(f"  Emails comerciais:   {progress.get('emails_comercial', 0)}")
    print(f"  Total de emails:     {total_emails}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    asyncio.run(main())
