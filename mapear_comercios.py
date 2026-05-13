"""
Mapeador de Comércios - Baixada Fluminense, RJ
===============================================
Busca comércios no Google Maps que NÃO possuem site,
gerando uma planilha de leads para prospecção.

Uso: python mapear_comercios.py
"""

import asyncio
import csv
import json
import os
import random
import re
import sys
from datetime import datetime
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent / "output" / "playwright"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Lista de cidades da Baixada Fluminense
CIDADES = [
    "Duque de Caxias, RJ",
    "Nova Iguaçu, RJ",
    "São João de Meriti, RJ",
    "Belford Roxo, RJ",
    "Nilópolis, RJ",
    "Mesquita, RJ",
    "Queimados, RJ",
    "Itaguaí, RJ",
    "Seropédica, RJ",
    "Paracambi, RJ",
    "Japeri, RJ",
]
MAX_RESULTS_PER_CATEGORY = 20  # Reduzido para ser mais rápido
DELAY_MIN = 2.0
DELAY_MAX = 5.0

CATEGORIAS = [
    "restaurante",
    "salão de beleza",
    "barbearia",
    "clínica médica",
    "oficina mecânica",
    "pet shop",
    "loja de roupas",
    "padaria",
    "pizzaria",
    "bar",
    "farmácia",
    "dentista",
    "advogado",
    "contador",
    "imobiliária",
    "academia",
    "lanchonete",
    "supermercado",
    "material de construção",
    "auto escola",
    "lavanderia",
    "floricultura",
    "ótica",
    "joalheria",
    "clínica veterinária",
    "estética",
    "loja de celulares",
    "loja de móveis",
    "papelaria",
    "loja de bicicleta",
    "confeitaria",
    "serralheria",
    "vidraçaria",
    "pintor",
    "eletricista",
    "encanador",
    "marcenaria",
    "escola de idiomas",
    "curso pré-vestibular",
    "estúdio de pilates",
]

PROGRESS_FILE = OUTPUT_DIR / "progresso.json"
EXISTING_CSV = OUTPUT_DIR / "todos_comercios.csv"


# ── Progress / Resume ──────────────────────────────────────────────

def load_nomes_existentes():
    """Carrega nomes+cidade do CSV existente para evitar duplicatas."""
    existentes = set()
    if EXISTING_CSV.exists():
        with open(EXISTING_CSV, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                nome = row.get("nome", "").strip().lower()
                cidade = row.get("cidade", "").strip().lower()
                if nome and cidade:
                    existentes.add(f"{nome}|{cidade}")
    return existentes

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"categorias_prontas": [], "categorias_em_andamento": {}, "comercios": []}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ── Export ──────────────────────────────────────────────────────────

def export_csv(businesses, filename):
    filepath = OUTPUT_DIR / filename
    fieldnames = [
        "nome", "endereco", "telefone", "whatsapp", "instagram",
        "email", "categoria", "tem_site", "url_site",
        "avaliacao", "num_avaliacoes", "cidade", "bairro", "link_maps",
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(businesses)
    return filepath


def export_excel(businesses, filename):
    """Exporta para .xlsx (requer openpyxl)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads sem Site"

        headers = [
            "Nome", "Endereço", "Telefone", "WhatsApp", "Instagram",
            "Email", "Categoria", "Tem Site?", "URL do Site",
            "Avaliação", "Nº Avaliações", "Cidade", "Bairro", "Link Maps",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, b in enumerate(businesses, 2):
            ws.cell(row=row_idx, column=1, value=b["nome"])
            ws.cell(row=row_idx, column=2, value=b["endereco"])
            ws.cell(row=row_idx, column=3, value=b["telefone"])
            ws.cell(row=row_idx, column=4, value=b.get("whatsapp", ""))
            ws.cell(row=row_idx, column=5, value=b.get("instagram", ""))
            ws.cell(row=row_idx, column=6, value=b.get("email", ""))
            ws.cell(row=row_idx, column=7, value=b["categoria"])
            ws.cell(row=row_idx, column=8, value="Sim" if b["tem_site"] else "NÃO")
            ws.cell(row=row_idx, column=9, value=b["url_site"])
            ws.cell(row=row_idx, column=10, value=b["avaliacao"])
            ws.cell(row=row_idx, column=11, value=b["num_avaliacoes"])
            ws.cell(row=row_idx, column=12, value=b.get("cidade", ""))
            ws.cell(row=row_idx, column=13, value=b.get("bairro", ""))
            ws.cell(row=row_idx, column=14, value=b.get("link_maps", ""))

            # Destaca leads sem site em amarelo
            if not b["tem_site"]:
                highlight = PatternFill("solid", fgColor="FEF08A")
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).fill = highlight

        # Ajusta largura das colunas
        widths = [35, 45, 18, 18, 30, 30, 25, 10, 35, 10, 12, 20, 22, 45]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        filepath = OUTPUT_DIR / filename
        wb.save(filepath)
        return filepath
    except ImportError:
        print("  [!] openpyxl não instalado — exportando CSV apenas")
        return None


# ── Google Maps Scraper ────────────────────────────────────────────

async def aceitar_cookies(page):
    """Clica no botão de aceitar cookies do Google, se aparecer."""
    for selector in [
        'button[aria-label*="Aceitar"]',
        'button[aria-label*="Accept"]',
        'button[aria-label*="accept all"]',
        'form:nth-of-type(2) button',
    ]:
        btn = page.locator(selector).first
        if await btn.count() > 0:
            try:
                await btn.click(timeout=2000)
                await asyncio.sleep(1)
                return True
            except Exception:
                pass
    return False


async def scroll_panel(page, max_scrolls=25):
    """Rola o painel de resultados para carregar mais comércios."""
    panel = page.locator('div[role="feed"]').first
    if await panel.count() == 0:
        return
    for _ in range(max_scrolls):
        await panel.evaluate("el => el.scrollBy(0, 800)")
        await asyncio.sleep(random.uniform(0.3, 0.8))
        fim = page.locator(
            'span:has-text("fim da lista"), span:has-text("end of the list")'
        )
        if await fim.count() > 0:
            break


async def extrair_detalhes(page, categoria, cidade=""):
    """Extrai dados de um comércio na página de detalhes aberta."""
    await page.wait_for_timeout(1500)

    dados = {
        "nome": "",
        "endereco": "",
        "telefone": "",
        "whatsapp": "",
        "instagram": "",
        "email": "",
        "categoria": categoria,
        "tem_site": False,
        "url_site": "",
        "avaliacao": "",
        "num_avaliacoes": "",
        "cidade": cidade,
        "bairro": "",
        "link_maps": "",
    }

    # Nome - mais seletivo para evitar pegar elementos errados
    for sel in ['h1.DUwDvf', 'h1[class*="fontHeadline"]', 'h1[class*="fontTitle"]']:
        el = page.locator(sel).first
        if await el.count() > 0:
            txt = (await el.inner_text()).strip()
            if txt and len(txt) > 3:  # Nome deve ter mais de 3 caracteres
                dados["nome"] = txt
                break

    # Validação: ignora nomes inválidos
    if not dados["nome"] or len(dados["nome"]) < 5:
        return None

    # Ignora entradas com nomes genéricos/inválidos
    nomes_invalidos = ["resultados", "todos", "categorias", "ver mais", "mostrar mais",
                       "próximo", "anterior", "fechar", "voltar", "menu", "pesquisa"]
    if dados["nome"].lower() in nomes_invalidos:
        return None

    # Avaliação
    rating_el = page.locator('div[role="img"][aria-label*="estrela"], div[role="img"][aria-label*="star"]').first
    if await rating_el.count() > 0:
        label = await rating_el.get_attribute("aria-label") or ""
        m = re.search(r"([\d,.]+)", label)
        if m:
            dados["avaliacao"] = m.group(1).replace(",", ".")

    # Nº avaliações
    reviews_el = page.locator('button[aria-label*="avaliação"], button[aria-label*="review"]').first
    if await reviews_el.count() > 0:
        txt = await reviews_el.inner_text()
        dados["num_avaliacoes"] = re.sub(r"\D", "", txt)

    # Endereço
    addr_el = page.locator(
        'button[data-item-id*="address"], button[aria-label*="Endereço"], '
        'button[aria-label*="Address"]'
    ).first
    if await addr_el.count() > 0:
        dados["endereco"] = (await addr_el.inner_text()).strip()
        # Remove caracteres especiais do início
        dados["endereco"] = re.sub(r'^[\W\u200b-\u200d]+', '', dados["endereco"])

    # Telefone
    phone_el = page.locator(
        'button[data-item-id*="phone:tel"], button[aria-label*="Telefone"], '
        'button[aria-label*="Phone"]'
    ).first
    if await phone_el.count() > 0:
        dados["telefone"] = (await phone_el.inner_text()).strip()
        # Remove caracteres especiais do início
        dados["telefone"] = re.sub(r'^[\W\u200b-\u200d]+', '', dados["telefone"])

    # Website
    web_el = page.locator(
        'a[data-item-id*="authority"], a[aria-label*="Site"], '
        'a[aria-label*="Website"], a[aria-label*="site oficial"]'
    ).first
    if await web_el.count() > 0:
        dados["tem_site"] = True
        dados["url_site"] = (await web_el.get_attribute("href")) or ""

    # WhatsApp
    whatsapp_el = page.locator(
        'a[href*="wa.me"], a[href*="whatsapp"], '
        'a[data-item-id*="whatsapp"], button[aria-label*="WhatsApp"], '
        'a[aria-label*="WhatsApp"]'
    ).first
    if await whatsapp_el.count() > 0:
        href = await whatsapp_el.get_attribute("href") or ""
        if "wa.me" in href:
            # Extrai o número do link wa.me
            wa_match = re.search(r"wa\.me/(\d+)", href)
            if wa_match:
                dados["whatsapp"] = wa_match.group(1)
        elif "whatsapp" in href.lower():
            dados["whatsapp"] = href
        else:
            dados["whatsapp"] = (await whatsapp_el.inner_text()).strip()

    # Se não achou WhatsApp dedicado, verifica se o telefone é celular (9° dígito)
    if not dados["whatsapp"] and dados["telefone"]:
        tel_limpo = re.sub(r"\D", "", dados["telefone"])
        # Celulares no Brasil têm 11 dígitos (DDD + 9 + 8 dígitos)
        if len(tel_limpo) == 11 or (len(tel_limpo) == 13 and tel_limpo.startswith("55")):
            dados["whatsapp"] = dados["telefone"]

    # Instagram
    insta_el = page.locator(
        'a[href*="instagram.com"], a[aria-label*="Instagram"]'
    ).first
    if await insta_el.count() > 0:
        insta_href = await insta_el.get_attribute("href") or ""
        if "instagram.com" in insta_href:
            # Limpa URL do Instagram
            insta_href = insta_href.rstrip("/")
            dados["instagram"] = insta_href
        else:
            dados["instagram"] = (await insta_el.inner_text()).strip()

    # Se o site for Instagram, captura como Instagram também
    if dados["url_site"] and "instagram.com" in dados["url_site"].lower():
        if not dados["instagram"]:
            dados["instagram"] = dados["url_site"].rstrip("/")
        dados["tem_site"] = False  # Instagram não é site próprio

    # Link do Google Maps (captura URL atual da página de detalhe)
    try:
        current_url = page.url
        if "/maps/" in current_url or "google" in current_url:
            dados["link_maps"] = current_url
    except Exception:
        pass

    # Bairro - extrair do endereço
    if dados["endereco"]:
        end = dados["endereco"]
        partes = end.split(" - ")
        if len(partes) >= 3:
            dados["bairro"] = partes[-2].strip().split(",")[0].strip()
        elif len(partes) == 2:
            bairro_cidade = partes[-1].strip()
            # Tenta separar bairro de cidade
            bc_parts = bairro_cidade.split(",")
            if len(bc_parts) >= 2:
                dados["bairro"] = bc_parts[0].strip()

    # Email - Tenta encontrar em vários locais
    email_selectors = [
        'a[href^="mailto:"]',
        'button[data-item-id*="email"]',
        'div[aria-label*="Email"]',
        'span:has-text("@")',
    ]

    for sel in email_selectors:
        email_el = page.locator(sel).first
        if await email_el.count() > 0:
            try:
                # Se for um link mailto:
                href = await email_el.get_attribute("href")
                if href and href.startswith("mailto:"):
                    dados["email"] = href.replace("mailto:", "").strip().split('?')[0]
                    break

                # Se for texto, tenta extrair email
                text = await email_el.inner_text()
                email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', text)
                if email_match:
                    dados["email"] = email_match.group(0)
                    break
            except Exception:
                pass

    return dados


async def buscar_categoria(page, categoria, cidade, start_index=0, progress=None, nomes_existentes=None):
    """Busca uma categoria no Google Maps e retorna lista de comércios."""
    query = f"{categoria} em {cidade}"
    url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}/"

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=45000)
    except Exception as e:
        print(f"  [!] Erro ao carregar: {e}")
        try:
            await page.reload(wait_until="commit", timeout=30000)
        except Exception:
            pass

    await asyncio.sleep(random.uniform(2, 3))
    await aceitar_cookies(page)
    await asyncio.sleep(2)

    # Rola para carregar mais
    await scroll_panel(page)
    await asyncio.sleep(1)

    # Captura todos os resultados
    items = page.locator('div[role="feed"] > div > div[jsaction], div[role="feed"] > div > a[jsaction]')
    total = await items.count()
    print(f"  > {total} resultados encontrados")

    comercios = []
    vistos = set()  # Controle de duplicatas nesta sessão

    # Carregar nomes já existentes no CSV para pular
    if nomes_existentes:
        for chave in nomes_existentes:
            if cidade.lower() in chave:
                vistos.add(chave)
    limite = min(total, MAX_RESULTS_PER_CATEGORY * 3)  # Tentar até 3x mais que o limite

    # Se está continuando de onde parou, carrega já vistos do progresso
    if start_index > 0:
        print(f"  > Continuando do indice {start_index}...")
        # Carrega duplicatas já vistas do progresso salvo
        for c in progress.get("comercios", []):
            if c.get("categoria") == categoria and c.get("cidade") == cidade:
                chave = f"{c['nome'].lower()}|{cidade}"
                vistos.add(chave)
        print(f"  > {len(vistos)} comércios já processados")

    # Loop até encontrar MAX_RESULTS únicos ou atingir limite
    i = start_index
    novos_encontrados = 0  # Conta só os novos (não os que já existiam no CSV)
    ja_existentes = len(vistos)

    while i < limite and novos_encontrados < MAX_RESULTS_PER_CATEGORY:
        # Recarrega os itens a cada iteração (o DOM pode mudar)
        items = page.locator('div[role="feed"] > div > div[jsaction], div[role="feed"] > div > a[jsaction]')
        total_itens = await items.count()

        if total_itens <= i:
            print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] Itens reduzidos, recarregando pagina...")
            await scroll_panel(page)
            await asyncio.sleep(1)
            items = page.locator('div[role="feed"] > div > div[jsaction], div[role="feed"] > div > a[jsaction]')
            total_itens = await items.count()
            if total_itens <= i:
                # Verifica se chegou ao fim da lista
                fim_lista = page.locator('span:has-text("fim da lista"), span:has-text("end of the list")')
                if await fim_lista.count() > 0:
                    print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] Fim da lista - Sem mais resultados")
                    break
                else:
                    print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] Nao ha mais itens disponiveis")
                    break

        # Verifica se chegou ao fim da lista antes de processar
        fim_lista = page.locator('span:has-text("fim da lista"), span:has-text("end of the list")')
        if await fim_lista.count() > 0 and i >= total_itens - 1:
            print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] Fim da lista atingido")
            break

        item = items.nth(i)

        # Verifica se o elemento está visível antes de tentar clicar
        try:
            if not await item.is_visible():
                print(f"    [{novos_encontrados+1}/{MAX_RESULTS_PER_CATEGORY}] Elemento não visível, pulando...")
                i += 1
                continue
        except:
            # Se não conseguir verificar visibilidade, tenta clicar mesmo assim
            pass

        # Tenta clicar com retry e maior timeout
        clicou = False
        for tentativa in range(3):
            try:
                await item.click(timeout=10000)
                clicou = True
                break
            except Exception as e:
                erro_str = str(e).lower()
                # Se o erro for "not visible", pula este elemento
                if 'not visible' in erro_str or 'element is not visible' in erro_str:
                    print(f"    [{novos_encontrados+1}/{MAX_RESULTS_PER_CATEGORY}] Elemento invisível, pulando...")
                    break
                elif tentativa < 2:
                    print(f"    [{novos_encontrados+1}/{MAX_RESULTS_PER_CATEGORY}] Retry clique {tentativa + 1}/3...")
                    await asyncio.sleep(1)
                    # Tenta scroll com menor timeout
                    try:
                        await item.scroll_into_view_if_needed(timeout=1000)
                    except:
                        pass
                    await asyncio.sleep(0.5)
                else:
                    print(f"    [{novos_encontrados+1}/{MAX_RESULTS_PER_CATEGORY}] Erro ao clicar: {str(e)[:60]}")
                    break

        if not clicou:
            i += 1
            continue

        await asyncio.sleep(random.uniform(1.8, 3.5))

        dados = await extrair_detalhes(page, categoria, cidade)
        if dados:
            # Verifica duplicata por nome + cidade
            chave = f"{dados['nome'].lower()}|{cidade.lower()}"
            if chave in vistos:
                print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] DUPLICATA - {dados['nome'][:40]}")
            else:
                vistos.add(chave)
                comercios.append(dados)
                novos_encontrados += 1
                tag = "COM site" if dados["tem_site"] else "SEM site"
                email_info = f" | Email: {dados['email']}" if dados['email'] else ""
                print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] {tag} - {dados['nome'][:40]}{email_info}")

                # Salva progresso incrementalmente
                if progress:
                    chave_progresso = f"{cidade}::{categoria}"
                    progress["categorias_em_andamento"][chave_progresso] = {"indice": i + 1, "total": limite}
                    progress["comercios"].extend(comercios)
                    comercios.clear()  # Limpa a lista local já que salvou no progress
                    save_progress(progress)

        # Volta para resultados - TENTATIVA 1: botão voltar
        voltou = False
        for _ in range(2):
            try:
                back = page.locator('button[aria-label*="Voltar"], button[aria-label*="Back"]').first
                if await back.count() > 0:
                    await back.click(timeout=3000)
                    await asyncio.sleep(0.5)
                    voltou = True
                    break
            except Exception:
                pass

            # TENTATIVA 2: go_back
            try:
                await page.go_back(timeout=3000)
                await asyncio.sleep(0.5)
                voltou = True
                break
            except Exception:
                pass

        if not voltou:
            # TENTATIVA 3: recarrega a URL de busca
            print(f"    [{novos_encontrados}/{MAX_RESULTS_PER_CATEGORY}] Navegacao travada, recarregando...")
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(2)
            await scroll_panel(page, max_scrolls=5)
            await asyncio.sleep(1)

        await asyncio.sleep(random.uniform(0.8, 1.5))
        i += 1

    # Mensagem final sobre a categoria
    if novos_encontrados >= MAX_RESULTS_PER_CATEGORY:
        print(f"  > Categoria completa: {novos_encontrados} novos resultados encontrados")
    else:
        print(f"  > Fim dos resultados: {novos_encontrados}/{MAX_RESULTS_PER_CATEGORY} novos ({ja_existentes} já existiam)")

    return comercios


# ── Main ────────────────────────────────────────────────────────────

async def main():
    # Força UTF-8 no Windows
    if sys.platform == "win32":
        import locale
        import codecs
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            pass

    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("[!] Playwright não encontrado. Instalando...")
        os.system(f"{sys.executable} -m pip install playwright")
        os.system(f"{sys.executable} -m playwright install chromium")
        from playwright.async_api import async_playwright

    print("=" * 60)
    print("  MAPEADOR DE COMÉRCIOS — BAIXADA FLUMINENSE, RJ")
    print("  Buscando comércios SEM site para prospecção")
    print(f"  {len(CIDADES)} cidades serão mapeadas")
    print("=" * 60)

    progress = load_progress()
    feitas = set(progress.get("categorias_prontas", []))
    em_andamento = progress.get("categorias_em_andamento", {})
    todos = progress.get("comercios", [])

    # Carregar nomes já existentes no CSV para evitar duplicatas
    nomes_existentes = load_nomes_existentes()
    if nomes_existentes:
        print(f"\n  {len(nomes_existentes)} comércios já existem no CSV (serão pulados)")

    # Formato antigo de progresso (cidade única) - migra para novo formato
    if not feitas and not em_andamento and not todos:
        print("\n[!] Progresso antigo detectado. Resetando para nova estrutura multi-cidade.")
        print("    Todas as cidades e categorias serão reprocessadas.\n")

    # Adiciona categorias em andamento às restantes (para continuar)
    restantes = [c for c in CATEGORIAS if c not in feitas]
    print(f"\nCategorias restantes: {len(restantes)}/{len(CATEGORIAS)}")
    print(f"Cidades a processar: {len(CIDADES)}")
    print(f"Comércios já mapeados: {len(todos)}")
    if em_andamento:
        print(f"Em andamento: {list(em_andamento.keys())}\n")

    if not restantes and not em_andamento:
        print("Todas as categorias já foram buscadas!")
        print("Delete output/playwright/progresso.json para recomeçar.\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1366, "height": 768},
            locale="pt-BR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
        )
        page = await ctx.new_page()

        # Abre o Google Maps e aceita cookies iniciais
        try:
            await page.goto("https://www.google.com/maps", wait_until="domcontentloaded", timeout=45000)
        except Exception as e:
            print(f"[!] Timeout ao carregar Maps. Tentando novamente... ({e})")
            await page.goto("https://www.google.com/maps", wait_until="commit", timeout=30000)
        await asyncio.sleep(3)
        await aceitar_cookies(page)
        await asyncio.sleep(1)

        # Itera sobre cada cidade e cada categoria
        total_tarefas = len(CIDADES) * len(CATEGORIAS)
        tarefa_atual = 0

        for cidade_idx, cidade in enumerate(CIDADES):
            print(f"\n{'='*60}")
            print(f"  CIDADE {cidade_idx+1}/{len(CIDADES)}: {cidade}")
            print(f"{'='*60}")

            for cat_idx, cat in enumerate(CATEGORIAS):
                tarefa_atual += 1

                # Pula categorias já concluídas
                chave_progresso = f"{cidade}::{cat}"
                if chave_progresso in feitas:
                    continue

                print(f"\n[{tarefa_atual}/{total_tarefas}] {cidade} - {cat}")

                # Verifica se já estava em andamento
                start_idx = 0
                if chave_progresso in em_andamento:
                    start_idx = em_andamento[chave_progresso].get("indice", 0)
                    print(f"  > Continuando do índice {start_idx}")

                try:
                    resultados = await buscar_categoria(page, cat, cidade, start_index=start_idx, progress=progress, nomes_existentes=nomes_existentes)

                    # Marca categoria como completa
                    if chave_progresso in em_andamento:
                        del em_andamento[chave_progresso]
                    feitas.add(chave_progresso)
                    progress["categorias_prontas"] = list(feitas)
                    progress["categorias_em_andamento"] = em_andamento
                    save_progress(progress)

                    # Conta sem site (do progresso atualizado)
                    sem = sum(1 for c in progress["comercios"] if c["categoria"] == cat and c.get("cidade") == cidade and not c["tem_site"])
                    todos_categoria = [c for c in progress["comercios"] if c["categoria"] == cat and c.get("cidade") == cidade]
                    print(f"  > {len(todos_categoria)} encontrados | {sem} sem site")

                except Exception as e:
                    print(f"  X Erro geral: {e}")
                    # Tenta navegar de volta ao Maps
                    try:
                        await page.goto("https://www.google.com/maps", wait_until="commit", timeout=15000)
                    except Exception:
                        pass
                    continue

                delay = random.uniform(DELAY_MIN, DELAY_MAX)
                print(f"  Aguardando {delay:.1f}s...")
                await asyncio.sleep(delay)

        await browser.close()

    # ── Exporta resultados ──────────────────────────────────────────
    # Recarrega o progresso para ter dados atualizados
    progress = load_progress()
    todos = progress.get("comercios", [])

    if not todos:
        print("\nNenhum comércio novo encontrado.")
        return

    # Mesclar com comércios já existentes no CSV
    existentes_csv = []
    if EXISTING_CSV.exists():
        with open(EXISTING_CSV, encoding="utf-8-sig") as f:
            existentes_csv = list(csv.DictReader(f))

    # Criar set de chaves dos novos para evitar duplicatas na mesclagem
    novos_nomes = set()
    for c in todos:
        novos_nomes.add(f"{c['nome'].lower().strip()}|{c.get('cidade', '').lower().strip()}")

    # Filtrar existentes que não estão nos novos
    mantidos = [c for c in existentes_csv
                if f"{c['nome'].lower().strip()}|{c.get('cidade', '').lower().strip()}" not in novos_nomes]

    todos_final = mantidos + todos
    print(f"\n  Mesclando: {len(mantidos)} existentes + {len(todos)} novos = {len(todos_final)} total")

    # Todos os comércios
    export_csv(todos_final, "todos_comercios.csv")
    print(f"\n✓ CSV completo: output/playwright/todos_comercios.csv")

    # Leads sem site
    sem_site = [c for c in todos_final if not c["tem_site"]]
    if sem_site:
        export_csv(sem_site, "leads_sem_site.csv")
        print(f"✓ CSV leads:    output/playwright/leads_sem_site.csv")

        xlsx = export_excel(sem_site, "leads_sem_site.xlsx")
        if xlsx:
            print(f"✓ Excel leads:  {xlsx}")

    # Resumo
    com_site = len(todos_final) - len(sem_site)
    taxa = (len(sem_site) / len(todos_final) * 100) if todos_final else 0
    print(f"\n{'='*50}")
    print(f"  RESUMO FINAL - BAIXADA FLUMINENSE")
    print(f"  Total mapeados:     {len(todos_final)}")
    print(f"  COM site:           {com_site}")
    print(f"  SEM site (leads):   {len(sem_site)}")
    print(f"  Taxa de prospecção: {taxa:.1f}%")
    print(f"{'='*50}")

    # Estatísticas por cidade
    print(f"\n  ESTATÍSTICAS POR CIDADE:")
    print(f"  {'-'*50}")
    for cidade in CIDADES:
        da_cidade = [c for c in todos_final if c.get("cidade") == cidade]
        sem_site_cidade = [c for c in da_cidade if not c["tem_site"]]
        if da_cidade:
            print(f"  {cidade.split(',')[0]:20s}: {len(da_cidade):4d} total | {len(sem_site_cidade):4d} sem site")
    print(f"{'='*50}")


if __name__ == "__main__":
    asyncio.run(main())
