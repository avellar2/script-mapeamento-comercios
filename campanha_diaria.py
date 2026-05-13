"""
Campanha Diária - Gerador de Planilha de Abordagem
====================================================
Lê leads_prospeccao.xlsx e gera uma planilha filtrada e organizada
para abordagem diária pelo WhatsApp.

Uso: python campanha_diaria.py [--arquivo CAMINHO] [--top N] [--hoje]
"""

import re
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] openpyxl não instalado. Instalando...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Diretórios ────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
PROSPECCAO_DIR = OUTPUT_DIR / "prospeccao"
CAMPANHAS_DIR = OUTPUT_DIR / "campanhas"


# ── Classificação por nicho ───────────────────────────────────────

NICHOS_ESTETICA = [
    "estética", "estetica", "salão de beleza", "salao de beleza",
    "manicure", "sobrancelha", "cílios", "cilios", "depilação",
    "depilacao", "spa", "clínica estética", "clinica estetica",
]

NICHOS_BARBEARIA = [
    "barbearia", "barbeiro", "salão masculino", "salao masculino",
]

NICHOS_COMIDA = [
    "restaurante", "marmitaria", "pizzaria", "lanchonete", "açaí",
    "açai", "confeitaria", "padaria", "churrascaria", "bar",
    "comida", "delivery", "hambúrguer", "hamburguer", "hot dog",
    "sushi", "japonesa", "acarajé", "acaraje", "petisco",
]

ABA_COMIDA = "Comida-Cardápio"

NICHOS_SERVICOS = [
    "assistência técnica", "assistencia tecnica", "refrigeração",
    "refrigeracao", "eletricista", "encanador", "manutenção",
    "manutencao", "conserto", "oficina mecânica", "oficina mecanica",
    "serralheria", "marcenaria", "pintor", "vidraçaria", "vidracaria",
]

NICHOS_IGREJA = [
    "igreja", "evento", "ministério", "ministerio",
]


# ── Mensagens WhatsApp por nicho ──────────────────────────────────

def gerar_mensagem_whatsapp(lead):
    nome = str(lead.get("nome", "")).strip()
    oferta = str(lead.get("oferta_sugerida", "")).strip()
    nicho = str(lead.get("nicho", "")).strip().lower()

    try:
        avaliacao = float(str(lead.get("nota_google", "0")).replace(",", "."))
    except (ValueError, TypeError):
        avaliacao = 0

    try:
        num_avaliacoes = int(re.sub(r"\D", "", str(lead.get("qtd_avaliacoes", "0"))))
    except (ValueError, TypeError):
        num_avaliacoes = 0

    # Abertura
    if avaliacao >= 4.0 and num_avaliacoes > 5:
        abertura = (
            f"Oi, tudo bem? Vi a {nome} no Google "
            f"e percebi que vocês têm boas avaliações."
        )
    else:
        abertura = f"Oi, tudo bem? Vi a {nome} no Google."

    # Corpo por nicho
    nicho_lower = nicho
    if any(n in nicho_lower for n in NICHOS_ESTETICA):
        corpo = (
            "Notei que as informações principais, como serviços, "
            "preços e agendamento, poderiam ficar mais fáceis de "
            "encontrar em uma página simples.\n\n"
            "Quando a cliente precisa procurar muito, ela pode acabar "
            "agendando em outro lugar.\n\n"
            "Eu trabalho criando páginas de agendamento para salões e "
            "clínicas de estética. Posso te mandar uma prévia visual de "
            "como ficaria?"
        )
    elif any(n in nicho_lower for n in NICHOS_BARBEARIA):
        corpo = (
            "Notei que os horários, preços e formas de agendar um corte "
            "podem ficar mais organizados em uma página simples.\n\n"
            "Quando o cliente quer agendar e não acha fácil, ele acaba "
            "procurando outra barbearia.\n\n"
            "Eu crio páginas de agendamento para barbearias. "
            "Posso te mandar uma prévia de como ficaria?"
        )
    elif any(n in nicho_lower for n in NICHOS_COMIDA):
        corpo = (
            "Notei que o cardápio e as formas de pedir poderiam ficar "
            "mais fáceis de acessar em uma página simples com o cardápio "
            "digital e botão direto para WhatsApp.\n\n"
            "Quando o cliente quer pedir e não encontra o cardápio fácil, "
            "ele pode acabar pedindo em outro lugar.\n\n"
            "Eu crio cardápios digitais para restaurantes e lanchonetes. "
            "Posso te mandar uma prévia de como ficaria?"
        )
    elif any(n in nicho_lower for n in NICHOS_SERVICOS):
        corpo = (
            "Notei que as informações de contato, serviços e orçamento "
            "podem ficar mais organizados em uma página simples.\n\n"
            "Quando o cliente precisa de um serviço rápido e não acha "
            "fácil como entrar em contato, ele pode ligar para outro "
            "profissional.\n\n"
            "Eu crio páginas com orçamento pelo WhatsApp para quem "
            "trabalha com serviços. Posso te mandar uma prévia?"
        )
    elif any(n in nicho_lower for n in NICHOS_IGREJA):
        corpo = (
            "Notei que a programação, endereço e informações de contato "
            "podem ficar mais organizados em uma página simples.\n\n"
            "Uma página ajuda as pessoas a encontrarem horários, "
            "endereço e inscrições facilmente.\n\n"
            "Eu crio páginas para eventos e igrejas. Posso te mandar "
            "uma prévia de como ficaria?"
        )
    else:
        corpo = (
            "Notei que as informações principais, como serviços, fotos, "
            "localização e contato, poderiam ficar mais organizadas em "
            "uma página simples.\n\n"
            "Quando o cliente precisa procurar muito, ele pode acabar "
            "desistindo ou chamando outro lugar.\n\n"
            "Eu trabalho criando páginas profissionais para negócios "
            "locais. Posso te mandar uma prévia visual de como ficaria?"
        )

    return f"{abertura}\n\n{corpo}"


def gerar_link_whatsapp(lead, mensagem):
    telefone = str(lead.get("whatsapp", "")).strip()
    if not telefone:
        telefone = str(lead.get("telefone", "")).strip()

    numeros = re.sub(r"\D", "", telefone)
    if not numeros:
        return ""

    if numeros.startswith("55") and len(numeros) >= 12:
        numero = numeros
    elif numeros.startswith("0"):
        numeros = numeros[1:]
        numero = "55" + numeros if len(numeros) in (10, 11) else ""
    elif len(numeros) == 11:
        numero = "55" + numeros
    elif len(numeros) == 10:
        numero = "55" + numeros
    else:
        return ""

    if not numero:
        return ""

    from urllib.parse import quote
    return f"https://wa.me/{numero}?text={quote(mensagem)}"


# ── Ação recomendada ──────────────────────────────────────────────

def gerar_acao_recomendada(lead):
    prioridade = str(lead.get("prioridade", "")).strip()
    site = str(lead.get("site", "")).strip()
    url_site = str(lead.get("url_site", "")).strip()

    if prioridade == "Alta":
        if site == "Não" or not url_site:
            return "Criar prévia visual e abordar pelo WhatsApp"
        # Site ruim ou incompleto (Facebook, Instagram, etc.)
        dominios_gratuitos = [
            "facebook.com", "instagram.com", "fb.me", "bit.ly",
            "wix.com", "wordpress.com", "blogspot.com",
        ]
        if any(d in url_site.lower() for d in dominios_gratuitos):
            return "Oferecer melhoria da página atual"
        return "Oferecer melhoria da página atual"

    if prioridade == "Média":
        return "Enviar modelo do nicho e testar interesse"

    return "Não abordar agora"


# ── Tipo de material ──────────────────────────────────────────────

def gerar_tipo_de_material(lead):
    nicho = str(lead.get("nicho", "")).strip().lower()

    if any(n in nicho for n in NICHOS_ESTETICA):
        return "Vídeo curto da demo de estética"
    if any(n in nicho for n in NICHOS_BARBEARIA):
        return "Vídeo curto da demo de barbearia"
    if any(n in nicho for n in NICHOS_COMIDA):
        return "Vídeo curto do cardápio digital"
    if any(n in nicho for n in NICHOS_SERVICOS):
        return "Print ou vídeo de mini site profissional"
    if any(n in nicho for n in NICHOS_IGREJA):
        return "Print ou vídeo de página para evento"
    return "Print do mini site vendedor"


# ── Classificação por aba ─────────────────────────────────────────

def classificar_aba(lead):
    nicho = str(lead.get("nicho", "")).strip().lower()
    if any(n in nicho for n in NICHOS_ESTETICA):
        return "Estética"
    if any(n in nicho for n in NICHOS_BARBEARIA):
        return "Barbearia"
    if any(n in nicho for n in NICHOS_COMIDA):
        return "Comida-Cardápio"
    if any(n in nicho for n in NICHOS_SERVICOS):
        return "Serviços Locais"
    return "Outros"


# ── Leitura do Excel de prospecção ────────────────────────────────

COLUNAS_PADRAO = [
    "nome", "nicho", "cidade", "bairro", "telefone", "whatsapp",
    "site", "url_site", "nota_google", "qtd_avaliacoes", "score",
    "prioridade", "oferta_sugerida", "motivo_prioridade",
    "mensagem_whatsapp", "link_whatsapp",
    "acao_recomendada", "tipo_de_material", "status",
    "data_abordagem", "data_followup", "observacoes",
]

# Mapeamento de colunas do Excel para nomes internos
MAPEAMENTO_EXCEL = {
    "Nome": "nome",
    "Nicho": "nicho",
    "Categoria": "nicho",
    "Cidade": "cidade",
    "Bairro": "bairro",
    "Telefone": "telefone",
    "WhatsApp": "whatsapp",
    "Instagram": "instagram",
    "Site": "site",
    "Tem Site?": "site",
    "URL Site": "url_site",
    "URL do Site": "url_site",
    "Nota": "nota_google",
    "Avaliação": "nota_google",
    "Avaliações": "qtd_avaliacoes",
    "Nº Avaliações": "qtd_avaliacoes",
    "Score": "score",
    "Prioridade": "prioridade",
    "Oferta Sugerida": "oferta_sugerida",
    "Motivo Prioridade": "motivo_prioridade",
    "Mensagem WhatsApp": "mensagem_whatsapp",
    "Link WhatsApp": "link_whatsapp",
    "Endereço": "endereco",
    "Email": "email",
    "Link Maps": "link_maps",
}


def ler_excel_prospeccao(caminho):
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["Leads"]

    # Lê cabeçalhos
    cabecalhos = []
    for cell in ws[1]:
        cabecalhos.append(str(cell.value).strip() if cell.value else "")

    # Mapeia colunas
    col_map = {}
    for idx, cab in enumerate(cabecalhos):
        nome_interno = MAPEAMENTO_EXCEL.get(cab, cab.lower().strip())
        col_map[nome_interno] = idx

    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lead = {}
        for nome_interno, idx in col_map.items():
            if idx < len(row):
                lead[nome_interno] = str(row[idx]) if row[idx] is not None else ""
            else:
                lead[nome_interno] = ""
        if lead.get("nome"):
            leads.append(lead)

    wb.close()
    return leads


def normalizar_lead(lead):
    """Normaliza campos para o formato da campanha."""
    saida = {}

    saida["nome"] = str(lead.get("nome", "")).strip()
    saida["nicho"] = str(lead.get("nicho", "")).strip()
    saida["cidade"] = str(lead.get("cidade", "")).strip()
    saida["bairro"] = str(lead.get("bairro", "")).strip()
    saida["telefone"] = str(lead.get("telefone", "")).strip()
    saida["whatsapp"] = str(lead.get("whatsapp", "")).strip()

    # Site: normaliza "True"/"False" para "Sim"/"Não"
    site_val = str(lead.get("site", "")).strip().lower()
    if site_val in ("true", "sim", "1", "yes"):
        saida["site"] = "Sim"
    else:
        saida["site"] = "Não"

    saida["url_site"] = str(lead.get("url_site", "")).strip()

    # Nota Google
    nota = str(lead.get("nota_google", lead.get("avaliacao", "0"))).strip()
    try:
        saida["nota_google"] = float(nota.replace(",", "."))
    except (ValueError, TypeError):
        saida["nota_google"] = 0

    # Qtd avaliações
    qtd = str(lead.get("qtd_avaliacoes", lead.get("num_avaliacoes", "0"))).strip()
    try:
        saida["qtd_avaliacoes"] = int(re.sub(r"\D", "", qtd))
    except (ValueError, TypeError):
        saida["qtd_avaliacoes"] = 0

    # Score
    try:
        saida["score"] = int(lead.get("score", 0))
    except (ValueError, TypeError):
        saida["score"] = 0

    saida["prioridade"] = str(lead.get("prioridade", "")).strip()
    saida["oferta_sugerida"] = str(lead.get("oferta_sugerida", "")).strip()
    saida["motivo_prioridade"] = str(lead.get("motivo_prioridade", "")).strip()

    # Mensagem WhatsApp melhorada
    saida["mensagem_whatsapp"] = gerar_mensagem_whatsapp(lead)

    # Link WhatsApp
    saida["link_whatsapp"] = gerar_link_whatsapp(lead, saida["mensagem_whatsapp"])

    # Ação recomendada
    saida["acao_recomendada"] = gerar_acao_recomendada(lead)

    # Tipo de material
    saida["tipo_de_material"] = gerar_tipo_de_material(lead)

    # Status
    saida["status"] = "novo"

    # Data de abordagem (vazio - preencher quando abordar)
    saida["data_abordagem"] = ""

    # Data de follow-up: 2 dias depois de hoje
    hoje = date.today()
    saida["data_followup"] = (hoje + timedelta(days=2)).strftime("%d/%m/%Y")

    # Observações (vazio)
    saida["observacoes"] = ""

    return saida


# ── Filtros ───────────────────────────────────────────────────────

def filtrar_leads_campanha(leads):
    """Filtra leads elegíveis para campanha diária."""
    filtrados = []
    for lead in leads:
        n = normalizar_lead(lead)

        # Prioridade Alta e score >= 70
        if n["prioridade"] != "Alta":
            continue
        if n["score"] < 70:
            continue

        # Deve ter telefone ou WhatsApp
        if not n["telefone"] and not n["whatsapp"]:
            continue

        # Sem site (quando a informação existe)
        if n["site"] == "Sim" and n["url_site"]:
            # Tem site profissional — pular
            dominios_gratuitos = [
                "facebook.com", "instagram.com", "fb.me", "bit.ly",
                "wix.com", "wordpress.com", "blogspot.com",
            ]
            if not any(d in n["url_site"].lower() for d in dominios_gratuitos):
                continue

        # Nicho bom para landing page
        nicho = n["nicho"].lower()
        nichos_bons = [
            "barbearia", "salão de beleza", "salao de beleza",
            "estética", "estetica", "manicure", "sobrancelha",
            "restaurante", "marmitaria", "pizzaria", "açai", "açaí",
            "lanchonete", "confeitaria", "padaria", "churrascaria", "bar",
            "advogado", "autônomo", "autonomo", "assistência técnica",
            "assistencia tecnica", "oficina mecânica", "oficina mecanica",
            "pet shop", "clínica veterinária", "clinica veterinaria",
            "auto escola", "academia", "estúdio de pilates", "estudio de pilates",
            "dentista", "clínica médica", "clinica medica",
            "eletricista", "encanador", "pintor", "marcenaria",
            "serralheria", "vidraçaria", "vidracaria",
            "material de construção", "igreja", "evento",
            "contador", "imobiliária", "imobiliaria",
            "loja de roupas", "loja de celulares", "loja de móveis",
            "loja de moveis", "floricultura", "ótica", "otica",
            "joalheria", "farmácia", "farmacia",
            "curso pré-vestibular", "escola de idiomas",
            "lavanderia", "papelaria", "loja de bicicleta",
        ]
        if not any(nb in nicho for nb in nichos_bons):
            continue

        filtrados.append(n)

    return filtrados


# ── Ordenação ─────────────────────────────────────────────────────

def ordenar_leads(leads):
    """Ordena por score, prioridade, avaliações, nota e WhatsApp."""
    def chave_ordenacao(lead):
        tem_whatsapp = 1 if lead.get("whatsapp") or lead.get("link_whatsapp") else 0
        return (
            -lead.get("score", 0),
            0 if lead.get("prioridade") == "Alta" else 1,
            -lead.get("qtd_avaliacoes", 0),
            -lead.get("nota_google", 0),
            -tem_whatsapp,
        )
    return sorted(leads, key=chave_ordenacao)


# ── Exportação Excel ──────────────────────────────────────────────

COR_ALTA = PatternFill("solid", fgColor="059669")
COR_MEDIA = PatternFill("solid", fgColor="D97706")
COR_BAIXA = PatternFill("solid", fgColor="DC2626")
COR_HEADER = PatternFill("solid", fgColor="1F2937")
COR_ROW_ALTA = PatternFill("solid", fgColor="ECFDF5")
COR_ROW_MEDIA = PatternFill("solid", fgColor="FFFBEB")
COR_ROW_BAIXA = PatternFill("solid", fgColor="FEF2F2")
COR_HOVER = PatternFill("solid", fgColor="F3F4F6")

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_ALTA = Font(bold=True, color="FFFFFF")
FONT_LINK = Font(color="2563EB", underline="single")
FONT_NORMAL = Font(size=10)
FONT_BOLD = Font(bold=True, size=10)

BORDER_THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

COLUNAS_EXCEL = [
    ("Nome", 38),
    ("Nicho", 22),
    ("Cidade", 20),
    ("Bairro", 22),
    ("Telefone", 18),
    ("WhatsApp", 18),
    ("Site", 8),
    ("Nota Google", 12),
    ("Avaliações", 12),
    ("Score", 8),
    ("Prioridade", 12),
    ("Oferta Sugerida", 24),
    ("Motivo Prioridade", 45),
    ("Ação Recomendada", 42),
    ("Tipo de Material", 34),
    ("Mensagem WhatsApp", 70),
    ("Link WhatsApp", 55),
    ("Status", 12),
    ("Data Abordagem", 16),
    ("Data Follow-up", 16),
    ("Observações", 30),
]

CAMPOS_INTERNOS = [
    "nome", "nicho", "cidade", "bairro", "telefone", "whatsapp",
    "site", "nota_google", "qtd_avaliacoes", "score", "prioridade",
    "oferta_sugerida", "motivo_prioridade", "acao_recomendada",
    "tipo_de_material", "mensagem_whatsapp", "link_whatsapp",
    "status", "data_abordagem", "data_followup", "observacoes",
]


def escrever_aba(ws, leads):
    """Escreve os leads em uma aba com formatação completa."""
    # Cabeçalho
    for col_idx, (nome, largura) in enumerate(COLUNAS_EXCEL, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = FONT_HEADER
        cell.fill = COR_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    # Dados
    for row_idx, lead in enumerate(leads, 2):
        prioridade = lead.get("prioridade", "")
        for col_idx, campo in enumerate(CAMPOS_INTERNOS, 1):
            valor = lead.get(campo, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = BORDER_THIN
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 13))

            # Cor da prioridade
            if col_idx == 11:  # Prioridade
                if prioridade == "Alta":
                    cell.fill = COR_ALTA
                    cell.font = FONT_ALTA
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif prioridade == "Média":
                    cell.fill = COR_MEDIA
                    cell.font = FONT_ALTA
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif prioridade == "Baixa":
                    cell.fill = COR_BAIXA
                    cell.font = FONT_ALTA
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Link WhatsApp clicável
            if col_idx == 17:  # Link WhatsApp
                link = str(valor).strip()
                if link.startswith("https://wa.me"):
                    cell.value = link
                    cell.font = FONT_LINK
                    cell.hyperlink = link

        # Cor de fundo da linha
        if prioridade == "Alta":
            row_fill = COR_ROW_ALTA
        elif prioridade == "Média":
            row_fill = COR_ROW_MEDIA
        else:
            row_fill = COR_ROW_BAIXA

        for col_idx in range(1, len(COLUNAS_EXCEL) + 1):
            if col_idx != 11:  # Não sobrescreve a coluna de prioridade
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # Filtros
    if leads:
        last_col = get_column_letter(len(COLUNAS_EXCEL))
        last_row = len(leads) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Congela primeira linha
    ws.freeze_panes = "A2"


def escrever_resumo(ws, leads_por_aba, total_geral):
    """Escreve a aba Resumo."""
    # Título
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="RESUMO DA CAMPANHA DIÁRIA")
    title.font = Font(bold=True, size=14, color="1F2937")
    title.alignment = Alignment(horizontal="center")

    hoje = date.today().strftime("%d/%m/%Y")
    ws.cell(row=2, column=1, value=f"Data: {hoje}").font = Font(size=11, color="6B7280")
    ws.merge_cells("A2:F2")

    row = 4
    ws.cell(row=row, column=1, value="Métrica").font = FONT_BOLD
    ws.cell(row=row, column=2, value="Valor").font = FONT_BOLD

    metricas = [
        ("Total de leads na campanha", total_geral),
        ("Leads com WhatsApp", sum(1 for l in leads_por_aba.get("Top 50", []) if l.get("whatsapp") or l.get("link_whatsapp"))),
        ("Leads sem site", sum(1 for l in leads_por_aba.get("Top 50", []) if l.get("site") == "Não")),
        ("Data de follow-up", (date.today() + timedelta(days=2)).strftime("%d/%m/%Y")),
    ]

    for label, valor in metricas:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=valor)

    # Por aba
    row += 2
    ws.cell(row=row, column=1, value="POR ABA").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    for col, header in enumerate(["Aba", "Qtd", "Alta", "Média", "Baixa", "Com WhatsApp"], 1):
        ws.cell(row=row, column=col, value=header).font = FONT_HEADER
        ws.cell(row=row, column=col).fill = COR_HEADER

    abas_ordem = ["Top 50", "Estética", "Barbearia", "Comida-Cardápio", "Serviços Locais", "Outros"]
    for nome_aba in abas_ordem:
        row += 1
        leads_aba = leads_por_aba.get(nome_aba, [])
        ws.cell(row=row, column=1, value=nome_aba)
        ws.cell(row=row, column=2, value=len(leads_aba))
        ws.cell(row=row, column=3, value=sum(1 for l in leads_aba if l.get("prioridade") == "Alta"))
        ws.cell(row=row, column=4, value=sum(1 for l in leads_aba if l.get("prioridade") == "Média"))
        ws.cell(row=row, column=5, value=sum(1 for l in leads_aba if l.get("prioridade") == "Baixa"))
        ws.cell(row=row, column=6, value=sum(1 for l in leads_aba if l.get("whatsapp") or l.get("link_whatsapp")))

    # Por tipo de material
    row += 2
    ws.cell(row=row, column=1, value="POR TIPO DE MATERIAL").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    for col, header in enumerate(["Tipo de Material", "Quantidade"], 1):
        ws.cell(row=row, column=col, value=header).font = FONT_HEADER
        ws.cell(row=row, column=col).fill = COR_HEADER

    todos_leads = []
    for leads_aba in leads_por_aba.values():
        todos_leads.extend(leads_aba)

    materiais = {}
    for l in todos_leads:
        mat = l.get("tipo_de_material", "")
        materiais[mat] = materiais.get(mat, 0) + 1

    for mat, qtd in sorted(materiais.items(), key=lambda x: x[1], reverse=True):
        row += 1
        ws.cell(row=row, column=1, value=mat)
        ws.cell(row=row, column=2, value=qtd)

    # Por oferta sugerida
    row += 2
    ws.cell(row=row, column=1, value="POR OFERTA SUGERIDA").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    for col, header in enumerate(["Oferta", "Quantidade"], 1):
        ws.cell(row=row, column=col, value=header).font = FONT_HEADER
        ws.cell(row=row, column=col).fill = COR_HEADER

    ofertas = {}
    for l in todos_leads:
        of = l.get("oferta_sugerida", "")
        ofertas[of] = ofertas.get(of, 0) + 1

    for of, qtd in sorted(ofertas.items(), key=lambda x: x[1], reverse=True):
        row += 1
        ws.cell(row=row, column=1, value=of)
        ws.cell(row=row, column=2, value=qtd)

    # Ações recomendadas
    row += 2
    ws.cell(row=row, column=1, value="POR AÇÃO RECOMENDADA").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    for col, header in enumerate(["Ação", "Quantidade"], 1):
        ws.cell(row=row, column=col, value=header).font = FONT_HEADER
        ws.cell(row=row, column=col).fill = COR_HEADER

    acoes = {}
    for l in todos_leads:
        a = l.get("acao_recomendada", "")
        acoes[a] = acoes.get(a, 0) + 1

    for a, qtd in sorted(acoes.items(), key=lambda x: x[1], reverse=True):
        row += 1
        ws.cell(row=row, column=1, value=a)
        ws.cell(row=row, column=2, value=qtd)

    # Larguras
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18


# ── Main ───────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Campanha Diária - Gerador de Planilha de Abordagem")
    parser.add_argument("--arquivo", help="Caminho para o Excel de prospecção (padrão: busca automaticamente)")
    parser.add_argument("--top", type=int, default=50, help="Quantidade de leads na aba Top 50 (padrão: 50)")
    parser.add_argument("--hoje", action="store_true", help="Inclui data de hoje na abordagem")
    args = parser.parse_args()

    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    print("=" * 60)
    print("  CAMPANHA DIÁRIA - GERADOR DE PLANILHA DE ABORDAGEM")
    print("=" * 60)

    # Busca arquivo de prospecção
    if args.arquivo:
        caminho = Path(args.arquivo)
        if not caminho.exists():
            print(f"[!] Arquivo não encontrado: {caminho}")
            sys.exit(1)
    else:
        caminho = PROSPECCAO_DIR / "leads_prospeccao.xlsx"
        if not caminho.exists():
            # Tenta buscar qualquer xlsx em prospeccao
            xlsx_files = list(PROSPECCAO_DIR.glob("*.xlsx"))
            if xlsx_files:
                caminho = xlsx_files[0]
            else:
                print("[!] Nenhum arquivo de prospecção encontrado.")
                print("    Execute prospectar_leads.py primeiro.")
                sys.exit(1)

    print(f"\n  Lendo: {caminho}")

    # Lê leads do Excel
    leads = ler_excel_prospeccao(caminho)
    print(f"  Leads carregados: {len(leads)}")

    # Filtra leads elegíveis
    leads_filtrados = filtrar_leads_campanha(leads)
    print(f"  Leads elegíveis (Alta + score≥70 + contato + sem site profissional): {len(leads_filtrados)}")

    if not leads_filtrados:
        print("\n[!] Nenhum lead elegível encontrado. Verifique os dados de prospecção.")
        sys.exit(0)

    # Se --hoje, preenche data_abordagem com hoje
    if args.hoje:
        hoje_str = date.today().strftime("%d/%m/%Y")
        for lead in leads_filtrados:
            lead["data_abordagem"] = hoje_str

    # Ordena
    leads_ordenados = ordenar_leads(leads_filtrados)

    # Classifica por aba
    leads_por_aba = {
        "Estética": [],
        "Barbearia": [],
        "Comida-Cardápio": [],
        "Serviços Locais": [],
        "Outros": [],
    }

    for lead in leads_ordenados:
        aba = classificar_aba(lead)
        leads_por_aba[aba].append(lead)

    # Top 50
    leads_por_aba["Top 50"] = leads_ordenados[:args.top]

    # Cria Excel
    CAMPANHAS_DIR.mkdir(parents=True, exist_ok=True)
    hoje_str = date.today().strftime("%Y-%m-%d")
    caminho_saida = CAMPANHAS_DIR / f"campanha_diaria_{hoje_str}.xlsx"

    wb = Workbook()
    # Remove a aba padrão
    wb.remove(wb.active)

    # Aba Top 50
    ws_top = wb.create_sheet("Top 50")
    escrever_aba(ws_top, leads_por_aba["Top 50"])
    print(f"\n  Top 50: {len(leads_por_aba['Top 50'])} leads")

    # Abas por nicho
    for nome_aba in ["Estética", "Barbearia", "Comida-Cardápio", "Serviços Locais", "Outros"]:
        ws = wb.create_sheet(nome_aba)
        escrever_aba(ws, leads_por_aba[nome_aba])
        print(f"  {nome_aba}: {len(leads_por_aba[nome_aba])} leads")

    # Aba Resumo
    ws_resumo = wb.create_sheet("Resumo")
    escrever_resumo(ws_resumo, leads_por_aba, len(leads_ordenados))
    print(f"  Resumo: OK")

    # Salva
    wb.save(caminho_saida)

    # Também salva como campanha_diaria.xlsx (última versão)
    caminho_padrao = CAMPANHAS_DIR / "campanha_diaria.xlsx"
    wb.save(caminho_padrao)

    print(f"\n{'='*60}")
    print(f"  ✓ Campanha salva: {caminho_saida}")
    print(f"  ✓ Atalho: {caminho_padrao}")
    print(f"\n  ROTINA RECOMENDADA:")
    print(f"  - Abra a planilha e vá na aba 'Top 50'")
    print(f"  - Selecione até 10 leads para abordar hoje")
    print(f"  - Clique no link WhatsApp de cada lead")
    print(f"  - Envie a mensagem personalizada")
    print(f"  - Anote a data na coluna 'Data Abordagem'")
    print(f"  - O follow-up está marcado para {(date.today() + timedelta(days=2)).strftime('%d/%m/%Y')}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()