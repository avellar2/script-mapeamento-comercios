"""
Prospecção de Leads - Gerador de Planilha Qualificada
=====================================================
Lê dados de comércios coletados e gera planilha de leads qualificados
com pontuação, prioridade, oferta sugerida e mensagem de WhatsApp.

Uso: python prospectar_leads.py [--arquivo CAMINHO] [--saida NOME]
"""

import csv
import re
import sys
from pathlib import Path
from urllib.parse import quote

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] openpyxl não instalado. Instalando...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Diretórios ────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
PLAYWRIGHT_DIR = OUTPUT_DIR / "playwright"
CONSOLIDADO_DIR = OUTPUT_DIR / "consolidado"
PROSPECCAO_DIR = OUTPUT_DIR / "prospeccao"


# ── Franquias conhecidas ──────────────────────────────────────────

FRANQUIAS_GRANDES = [
    "mcdonald", "burger king", "subway", "habib", "bob's", "kfc",
    "pizza hut", "domino", "starbucks", "giraffas", "spoleto",
    "outback", "applebee", "wendy", "taco bell", "baskin robbins",
    "cold stone", "nutrella", "camarao", "madero", "supermercado zaffari",
    "pao de acucar", "extra hiper", "carrefour", "assaí atacadista",
    "atacadao", "big", "mercado livre", "magazine luiza", "americanas",
    "casas bahia", "pontofrio", "extra", "renner", "riachuelo",
    "c&a", "marisa", "arezzo", "track&field", "vivolo",
    "smartfit", "bluefit", "bodytech", "formula academia",
    "odontoprev", "oral uni", "sorriso", "dentsply",
    "cabana", "farroupilha", "churrascaria rodizio",
    "porcao", "fogo de chao", "chama gaucho",
    "ifood", "rappi", "99food", "ubereats",
]

NICHOS_BONS_LANDING = [
    "barbearia", "salao de beleza", "salão de beleza", "estetica", "estética",
    "manicure", "sobrancelha", "restaurante", "marmitaria", "pizzaria",
    "açai", "açaí", "lanchonete", "confeitaria", "padaria", "churrascaria",
    "bar", "igreja", "evento", "advogado", "autônomo", "autonomo",
    "assistencia tecnica", "assistência técnica", "oficina mecanica",
    "oficina mecânica", "pet shop", "clinica veterinaria", "clínica veterinária",
    "auto escola", "academia", "estúdio de pilates", "estudio de pilates",
    "dentista", "clinica medica", "clínica médica",
]

OFERTA_POR_NICHO = {
    # Página de Agendamento
    "barbearia": "Página de Agendamento",
    "salão de beleza": "Página de Agendamento",
    "salao de beleza": "Página de Agendamento",
    "estética": "Página de Agendamento",
    "estetica": "Página de Agendamento",
    "manicure": "Página de Agendamento",
    "sobrancelha": "Página de Agendamento",
    "clínica médica": "Página de Agendamento",
    "clinica medica": "Página de Agendamento",
    "estúdio de pilates": "Página de Agendamento",
    "estudio de pilates": "Página de Agendamento",
    "academia": "Página de Agendamento",
    # Cardápio Digital
    "restaurante": "Cardápio Digital",
    "marmitaria": "Cardápio Digital",
    "pizzaria": "Cardápio Digital",
    "açai": "Cardápio Digital",
    "açaí": "Cardápio Digital",
    "lanchonete": "Cardápio Digital",
    "confeitaria": "Cardápio Digital",
    "padaria": "Cardápio Digital",
    "churrascaria": "Cardápio Digital",
    "bar": "Cardápio Digital",
    # Página de Evento
    "igreja": "Página de Evento",
    "evento": "Página de Evento",
    # Mini Site Profissional
    "advogado": "Mini Site Profissional",
    "autônomo": "Mini Site Profissional",
    "autonomo": "Mini Site Profissional",
    "assistência técnica": "Mini Site Profissional",
    "assistencia tecnica": "Mini Site Profissional",
    "contador": "Mini Site Profissional",
    "eletricista": "Mini Site Profissional",
    "encanador": "Mini Site Profissional",
    "pintor": "Mini Site Profissional",
    "marcenaria": "Mini Site Profissional",
    "serralheria": "Mini Site Profissional",
    "vidraçaria": "Mini Site Profissional",
    "material de construção": "Mini Site Profissional",
    # Mini Site Profissional - serviços
    "oficina mecânica": "Mini Site Profissional",
    "oficina mecanica": "Mini Site Profissional",
    "dentista": "Mini Site Profissional",
    "pet shop": "Mini Site Profissional",
    "clínica veterinária": "Mini Site Profissional",
    "clinica veterinaria": "Mini Site Profissional",
    "auto escola": "Mini Site Profissional",
    "escola de idiomas": "Mini Site Profissional",
    "curso pré-vestibular": "Mini Site Profissional",
    # Página de Agendamento - mais serviços
    "floricultura": "Mini Site Vendedor",
    "ótica": "Mini Site Vendedor",
    "otica": "Mini Site Vendedor",
    "joalheria": "Mini Site Vendedor",
    "loja de roupas": "Mini Site Vendedor",
    "loja de celulares": "Mini Site Vendedor",
    "loja de móveis": "Mini Site Vendedor",
    "loja de moveis": "Mini Site Vendedor",
    "loja de bicicleta": "Mini Site Vendedor",
    "papelaria": "Mini Site Vendedor",
    "farmácia": "Mini Site Vendedor",
    "farmacia": "Mini Site Vendedor",
    "supermercado": "Mini Site Vendedor",
    "lavanderia": "Mini Site Vendedor",
    "imobiliária": "Mini Site Vendedor",
    "imobiliaria": "Mini Site Vendedor",
}


# ── Funções de Score ──────────────────────────────────────────────

def eh_franquia(nome):
    nome_lower = nome.lower()
    for f in FRANQUIAS_GRANDES:
        if f in nome_lower:
            return True
    return False


def site_profissional(url_site):
    if not url_site:
        return False
    url = url_site.lower().strip()
    dominios_gratuitos = [
        "facebook.com", "instagram.com", "fb.me", "bit.ly", "tinyurl.com",
        "wix.com", "wordpress.com", "blogspot.com", "tumblr.com",
        "google.com/maps", "goo.gl/maps", "maps.app.goo.gl",
    ]
    for d in dominios_gratuitos:
        if d in url:
            return False
    if url.startswith("http") and "." in url:
        return True
    return False


def calcular_score(lead):
    score = 0

    # +30 se não tiver site
    tem_site = str(lead.get("tem_site", "")).strip().lower()
    if tem_site in ("false", "f", "0", "", "não", "nao", "n"):
        score += 30

    # +20 se tiver telefone ou WhatsApp
    telefone = str(lead.get("telefone", "")).strip()
    whatsapp = str(lead.get("whatsapp", "")).strip()
    if telefone or whatsapp:
        score += 20

    # +15 se tiver nota >= 4.0
    try:
        avaliacao = float(str(lead.get("avaliacao", "0")).replace(",", "."))
        if avaliacao >= 4.0:
            score += 15
    except (ValueError, TypeError):
        pass

    # +15 se tiver mais de 20 avaliações
    try:
        num_avaliacoes = int(re.sub(r"\D", "", str(lead.get("num_avaliacoes", "0"))))
        if num_avaliacoes > 20:
            score += 15
    except (ValueError, TypeError):
        pass

    # +10 se tiver Instagram
    instagram = str(lead.get("instagram", "")).strip()
    if instagram:
        score += 10

    # +10 se for nicho bom para landing page
    nicho = str(lead.get("categoria", "")).strip().lower()
    if any(n in nicho for n in NICHOS_BONS_LANDING):
        score += 10

    # -30 se for franquia grande
    nome = str(lead.get("nome", "")).strip()
    if eh_franquia(nome):
        score -= 30

    # -20 se já tiver site aparentemente profissional
    url_site = str(lead.get("url_site", "")).strip()
    if site_profissional(url_site):
        score -= 20

    return max(0, min(100, score))


def classificar_prioridade(score):
    if score >= 70:
        return "Alta"
    elif score >= 40:
        return "Média"
    return "Baixa"


def gerar_motivo_prioridade(lead, score):
    motivos = []

    tem_site = str(lead.get("tem_site", "")).strip().lower()
    telefone = str(lead.get("telefone", "")).strip()
    whatsapp = str(lead.get("whatsapp", "")).strip()
    instagram = str(lead.get("instagram", "")).strip()
    nome = str(lead.get("nome", "")).strip()
    url_site = str(lead.get("url_site", "")).strip()
    nicho = str(lead.get("categoria", "")).strip().lower()

    try:
        avaliacao = float(str(lead.get("avaliacao", "0")).replace(",", "."))
    except (ValueError, TypeError):
        avaliacao = 0

    try:
        num_avaliacoes = int(re.sub(r"\D", "", str(lead.get("num_avaliacoes", "0"))))
    except (ValueError, TypeError):
        num_avaliacoes = 0

    # Pontos positivos
    if tem_site in ("false", "f", "0", "", "não", "nao", "n"):
        motivos.append("sem site")

    if telefone or whatsapp:
        motivos.append("tem telefone/WhatsApp")

    if avaliacao >= 4.0:
        motivos.append(f"boa avaliação ({avaliacao:.1f})")

    if num_avaliacoes > 20:
        motivos.append(f"{num_avaliacoes} avaliações")

    if instagram:
        motivos.append("tem Instagram")

    if any(n in nicho for n in NICHOS_BONS_LANDING):
        motivos.append("nicho promissor")

    # Pontos negativos
    if eh_franquia(nome):
        motivos.append("franquia grande")

    if site_profissional(url_site):
        motivos.append("site profissional")

    if score < 40:
        if not motivos:
            motivos.append("lead fraco sem diferenciais")
    elif score >= 70:
        if not motivos:
            motivos.append("lead qualificado")

    return " | ".join(motivos) if motivos else "lead sem dados suficientes"


def gerar_oferta_sugerida(lead):
    nicho = str(lead.get("categoria", "")).strip().lower()
    if nicho in OFERTA_POR_NICHO:
        return OFERTA_POR_NICHO[nicho]
    return "Mini Site Vendedor"


def gerar_mensagem_whatsapp(lead):
    nome = str(lead.get("nome", "")).strip()
    oferta = gerar_oferta_sugerida(lead)
    tem_site = str(lead.get("tem_site", "")).strip().lower()
    telefone = str(lead.get("telefone", "")).strip()
    whatsapp = str(lead.get("whatsapp", "")).strip()

    try:
        avaliacao = float(str(lead.get("avaliacao", "0")).replace(",", "."))
    except (ValueError, TypeError):
        avaliacao = 0

    try:
        num_avaliacoes = int(re.sub(r"\D", "", str(lead.get("num_avaliacoes", "0"))))
    except (ValueError, TypeError):
        num_avaliacoes = 0

    tem_contato = bool(telefone or whatsapp)

    # Constrói a mensagem com base no perfil do lead
    partes = [f"Oi, tudo bem? Vi a {nome} no Google"]

    if avaliacao >= 4.0 and num_avaliacoes > 5:
        partes.append(f"e percebi que vocês têm boas avaliações ({avaliacao:.1f} estrelas)")
    else:
        partes.append("e notei algo que pode estar afastando clientes")

    sem_site = tem_site in ("false", "f", "0", "", "não", "nao", "n")
    if sem_site:
        partes.append(
            "mas ainda não encontrei uma página simples com serviços, fotos, "
            "localização e botão direto para WhatsApp"
        )
    else:
        partes.append(
            "mas a página atual parece não estar trazendo os resultados que vocês merecem"
        )

    partes.append(
        "Quando o cliente precisa procurar muito essas informações, "
        "ele pode acabar chamando outro lugar"
    )

    if oferta == "Cardápio Digital":
        partes.append(f"Eu crio {oferta.lower()}s profissionais para negócios como o seu")
    elif oferta == "Página de Agendamento":
        partes.append(f"Eu crio {oferta.lower()}s profissionais para negócios locais")
    else:
        partes.append("Eu crio páginas profissionais para negócios locais")

    partes.append("Posso te mandar uma prévia visual de como ficaria?")

    return " ".join(partes)


def limpar_telefone(telefone):
    if not telefone:
        return ""
    numeros = re.sub(r"\D", "", telefone)
    if not numeros:
        return ""
    # Brasil: adiciona 55 se não tiver
    if numeros.startswith("55") and len(numeros) >= 12:
        return numeros
    if numeros.startswith("0"):
        numeros = numeros[1:]
    if len(numeros) == 11:
        numeros = "55" + numeros
    elif len(numeros) == 10:
        numeros = "55" + numeros
    elif len(numeros) == 9:
        # Somente celular sem DDD — não dá pra usar
        return ""
    return numeros


def gerar_link_whatsapp(lead, mensagem):
    telefone = str(lead.get("whatsapp", "")).strip()
    if not telefone:
        telefone = str(lead.get("telefone", "")).strip()

    numero = limpar_telefone(telefone)
    if not numero:
        return ""

    msg_codificada = quote(mensagem)
    return f"https://wa.me/{numero}?text={msg_codificada}"


def extrair_bairro(endereco, cidade=""):
    if not endereco:
        return ""
    # Padrões comuns em endereços brasileiros:
    # "Rua Nome, 123 - Bairro, Cidade - RJ, CEP"
    # "Avenida Nome - Bairro, Cidade - RJ"
    # Tenta extrair o bairro entre o primeiro hífen e a cidade
    partes = endereco.split(" - ")
    if len(partes) >= 3:
        # O bairro geralmente está na segunda parte
        bairro = partes[1].strip()
        # Remove cidade se apareceu junto
        if cidade and cidade.split(",")[0].strip() in bairro:
            bairro = bairro.replace(cidade.split(",")[0].strip(), "").strip(" ,")
        return bairro
    if len(partes) == 2:
        bairro = partes[1].split(",")[0].strip()
        return bairro
    # Se não tem hífen, tenta vírgula
    partes_virgula = endereco.split(",")
    if len(partes_virgula) >= 3:
        return partes_virgula[1].strip()
    return ""


# ── Normalização de Dados ─────────────────────────────────────────

MAPEAMENTO_COLUNAS = {
    # Nome
    "nome": "nome",
    "name": "nome",
    "nome_do_comercio": "nome",
    # Endereço
    "endereco": "endereco",
    "endereço": "endereco",
    "address": "endereco",
    # Telefone
    "telefone": "telefone",
    "phone": "telefone",
    "phone_number": "telefone",
    # WhatsApp
    "whatsapp": "whatsapp",
    # Instagram
    "instagram": "instagram",
    # Cidade
    "cidade": "cidade",
    "city": "cidade",
    # Bairro
    "bairro": "bairro",
    "neighborhood": "bairro",
    # Nicho/Categoria
    "categoria": "categoria",
    "nicho": "categoria",
    "category": "categoria",
    # Site
    "tem_site": "tem_site",
    "has_website": "tem_site",
    "url_site": "url_site",
    "website": "url_site",
    "site": "url_site",
    # Avaliação
    "avaliacao": "avaliacao",
    "avaliação": "avaliacao",
    "rating": "avaliacao",
    "nota": "avaliacao",
    # Número de avaliações
    "num_avaliacoes": "num_avaliacoes",
    "n_avaliacoes": "num_avaliacoes",
    "review_count": "num_avaliacoes",
    "reviews": "num_avaliacoes",
    # Email
    "email": "email",
    "email_cnpj": "email",
    "email_comercial": "email",
    # Link Maps
    "link_maps": "link_maps",
    "google_maps_url": "link_maps",
}


def normalizar_lead(row):
    lead = {}
    for col_original, valor in row.items():
        col = col_original.strip().lower()
        col_normalizada = MAPEAMENTO_COLUNAS.get(col, col)
        if col_normalizada not in lead:
            lead[col_normalizada] = valor.strip() if isinstance(valor, str) else valor

    # Garante campos obrigatórios
    lead.setdefault("nome", "")
    lead.setdefault("endereco", "")
    lead.setdefault("telefone", "")
    lead.setdefault("whatsapp", "")
    lead.setdefault("instagram", "")
    lead.setdefault("cidade", "")
    lead.setdefault("bairro", "")
    lead.setdefault("categoria", "")
    lead.setdefault("tem_site", "")
    lead.setdefault("url_site", "")
    lead.setdefault("avaliacao", "")
    lead.setdefault("num_avaliacoes", "")
    lead.setdefault("email", "")
    lead.setdefault("link_maps", "")

    # Extrai bairro do endereço se não estiver preenchido
    if not lead.get("bairro") and lead.get("endereco"):
        lead["bairro"] = extrair_bairro(lead["endereco"], lead.get("cidade", ""))

    # Limpa cidade (remove ", RJ" etc.)
    cidade = str(lead.get("cidade", "")).strip()
    if ", RJ" in cidade:
        lead["cidade"] = cidade.replace(", RJ", "").replace(",RJ", "").strip()

    # Normaliza tem_site
    tem_site = str(lead.get("tem_site", "")).strip().lower()
    if tem_site in ("true", "t", "1", "sim", "yes"):
        lead["tem_site"] = "True"
    else:
        lead["tem_site"] = "False"

    return lead


# ── Carregamento de Dados ─────────────────────────────────────────

def encontrar_arquivo_dados(caminho=None):
    if caminho:
        p = Path(caminho)
        if p.exists():
            return p
        print(f"[!] Arquivo não encontrado: {caminho}")
        sys.exit(1)

    # Prioridade: consolidado > playwright
    prioridades = [
        CONSOLIDADO_DIR / "base_emails_consolidada.csv",
        CONSOLIDADO_DIR / "base_emails_pronta_contato_refinada.csv",
        PLAYWRIGHT_DIR / "todos_comercios.csv",
    ]

    for p in prioridades:
        if p.exists():
            return p

    print("[!] Nenhum arquivo de dados encontrado em output/")
    print("    Execute mapear_comercios.py primeiro.")
    sys.exit(1)


def carregar_dados(caminho):
    leads = []
    with open(caminho, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            lead = normalizar_lead(row)
            if lead.get("nome"):
                leads.append(lead)

    # Remove duplicatas por nome|cidade
    vistos = set()
    unicos = []
    for lead in leads:
        chave = f"{lead['nome'].lower().strip()}|{lead.get('cidade', '').lower().strip()}"
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(lead)

    return unicos


# ── Exportação Excel ──────────────────────────────────────────────

CORES_PRIORIDADE = {
    "Alta": PatternFill("solid", fgColor="059669"),   # Verde
    "Média": PatternFill("solid", fgColor="D97706"),  # Amarelo
    "Baixa": PatternFill("solid", fgColor="DC2626"),  # Vermelho
}

FONTES_PRIORIDADE = {
    "Alta": Font(bold=True, color="FFFFFF"),
    "Média": Font(bold=True, color="FFFFFF"),
    "Baixa": Font(bold=True, color="FFFFFF"),
}


def exportar_excel(leads, caminho_saida):
    PROSPECCAO_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Aba Leads ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leads"

    colunas = [
        ("Prioridade", 14),
        ("Score", 8),
        ("Nome", 38),
        ("Nicho", 22),
        ("Cidade", 20),
        ("Bairro", 22),
        ("Telefone", 18),
        ("WhatsApp", 18),
        ("Instagram", 28),
        ("Site", 8),
        ("URL Site", 35),
        ("Nota", 8),
        ("Avaliações", 12),
        ("Endereço", 45),
        ("Link Maps", 45),
        ("Oferta Sugerida", 24),
        ("Motivo Prioridade", 50),
        ("Mensagem WhatsApp", 80),
        ("Link WhatsApp", 120),
        ("Email", 30),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, (nome, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    # Ordena por score descendente (alta prioridade no topo)
    leads_ordenados = sorted(leads, key=lambda l: l.get("score", 0), reverse=True)

    for row_idx, lead in enumerate(leads_ordenados, 2):
        prioridade = lead.get("prioridade", "")
        score = lead.get("score", 0)

        valores = [
            prioridade,
            score,
            lead.get("nome", ""),
            lead.get("categoria", ""),
            lead.get("cidade", ""),
            lead.get("bairro", ""),
            lead.get("telefone", ""),
            lead.get("whatsapp", ""),
            lead.get("instagram", ""),
            lead.get("tem_site", ""),
            lead.get("url_site", ""),
            lead.get("avaliacao", ""),
            lead.get("num_avaliacoes", ""),
            lead.get("endereco", ""),
            lead.get("link_maps", ""),
            lead.get("oferta_sugerida", ""),
            lead.get("motivo_prioridade", ""),
            lead.get("mensagem_whatsapp", ""),
            lead.get("link_whatsapp", ""),
            lead.get("email", ""),
        ]

        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 16))

            # Cor da prioridade na coluna 1
            if col_idx == 1 and prioridade in CORES_PRIORIDADE:
                cell.fill = CORES_PRIORIDADE[prioridade]
                cell.font = FONTES_PRIORIDADE[prioridade]
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Linha com cor de fundo sutil por prioridade
        if prioridade == "Alta":
            row_fill = PatternFill("solid", fgColor="ECFDF5")
        elif prioridade == "Média":
            row_fill = PatternFill("solid", fgColor="FFFBEB")
        else:
            row_fill = PatternFill("solid", fgColor="FEF2F2")

        for col_idx in range(2, len(colunas) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # Filtros automáticos
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(leads_ordenados) + 1}"

    # Congela cabeçalho
    ws.freeze_panes = "A2"

    # ── Aba Resumo ────────────────────────────────────────────
    ws_resumo = wb.create_sheet("Resumo")

    total = len(leads_ordenados)
    alta = [l for l in leads_ordenados if l.get("prioridade") == "Alta"]
    media = [l for l in leads_ordenados if l.get("prioridade") == "Média"]
    baixa = [l for l in leads_ordenados if l.get("prioridade") == "Baixa"]

    media_score = sum(l.get("score", 0) for l in leads_ordenados) / total if total else 0

    resumo_header_font = Font(bold=True, color="FFFFFF", size=12)
    resumo_header_fill = PatternFill("solid", fgColor="1F2937")
    resumo_label_font = Font(bold=True, size=11)
    resumo_valor_font = Font(size=11)
    resumo_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Título
    ws_resumo.merge_cells("A1:D1")
    title_cell = ws_resumo.cell(row=1, column=1, value="RESUMO DA PROSPECÇÃO")
    title_cell.font = Font(bold=True, size=14, color="1F2937")
    title_cell.alignment = Alignment(horizontal="center")

    # Estatísticas gerais
    stats = [
        ("Total de Leads", total),
        ("Leads Alta Prioridade", len(alta)),
        ("Leads Média Prioridade", len(media)),
        ("Leads Baixa Prioridade", len(baixa)),
        ("Score Médio", f"{media_score:.1f}"),
        ("Score Médio (Alta)", f"{sum(l['score'] for l in alta)/len(alta):.1f}" if alta else "0"),
        ("Score Médio (Média)", f"{sum(l['score'] for l in media)/len(media):.1f}" if media else "0"),
        ("Score Médio (Baixa)", f"{sum(l['score'] for l in baixa)/len(baixa):.1f}" if baixa else "0"),
    ]

    row = 3
    ws_resumo.cell(row=row, column=1, value="Métrica").font = resumo_header_font
    ws_resumo.cell(row=row, column=1).fill = resumo_header_fill
    ws_resumo.cell(row=row, column=2, value="Valor").font = resumo_header_font
    ws_resumo.cell(row=row, column=2).fill = resumo_header_fill

    for label, valor in stats:
        row += 1
        ws_resumo.cell(row=row, column=1, value=label).font = resumo_label_font
        ws_resumo.cell(row=row, column=2, value=valor).font = resumo_valor_font

    # Por cidade
    row += 2
    ws_resumo.cell(row=row, column=1, value="POR CIDADE").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    ws_resumo.cell(row=row, column=1, value="Cidade").font = resumo_header_font
    ws_resumo.cell(row=row, column=1).fill = resumo_header_fill
    ws_resumo.cell(row=row, column=2, value="Total").font = resumo_header_font
    ws_resumo.cell(row=row, column=2).fill = resumo_header_fill
    ws_resumo.cell(row=row, column=3, value="Alta").font = resumo_header_font
    ws_resumo.cell(row=row, column=3).fill = resumo_header_fill
    ws_resumo.cell(row=row, column=4, value="Média").font = resumo_header_font
    ws_resumo.cell(row=row, column=4).fill = resumo_header_fill
    ws_resumo.cell(row=row, column=5, value="Baixa").font = resumo_header_font
    ws_resumo.cell(row=row, column=5).fill = resumo_header_fill

    cidades = sorted(set(l.get("cidade", "") for l in leads_ordenados if l.get("cidade")))
    for cidade in cidades:
        row += 1
        da_cidade = [l for l in leads_ordenados if l.get("cidade") == cidade]
        ws_resumo.cell(row=row, column=1, value=cidade)
        ws_resumo.cell(row=row, column=2, value=len(da_cidade))
        ws_resumo.cell(row=row, column=3, value=len([l for l in da_cidade if l.get("prioridade") == "Alta"]))
        ws_resumo.cell(row=row, column=4, value=len([l for l in da_cidade if l.get("prioridade") == "Média"]))
        ws_resumo.cell(row=row, column=5, value=len([l for l in da_cidade if l.get("prioridade") == "Baixa"]))

    # Por nicho
    row += 2
    ws_resumo.cell(row=row, column=1, value="POR NICHO").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    for col, header in enumerate(["Nicho", "Total", "Alta", "Média", "Baixa", "Oferta Sugerida"], 1):
        ws_resumo.cell(row=row, column=col, value=header).font = resumo_header_font
        ws_resumo.cell(row=row, column=col).fill = resumo_header_fill

    nichos = sorted(set(l.get("categoria", "") for l in leads_ordenados if l.get("categoria")))
    for nicho in nichos:
        row += 1
        do_nicho = [l for l in leads_ordenados if l.get("categoria") == nicho]
        ws_resumo.cell(row=row, column=1, value=nicho)
        ws_resumo.cell(row=row, column=2, value=len(do_nicho))
        ws_resumo.cell(row=row, column=3, value=len([l for l in do_nicho if l.get("prioridade") == "Alta"]))
        ws_resumo.cell(row=row, column=4, value=len([l for l in do_nicho if l.get("prioridade") == "Média"]))
        ws_resumo.cell(row=row, column=5, value=len([l for l in do_nicho if l.get("prioridade") == "Baixa"]))
        ws_resumo.cell(row=row, column=6, value=gerar_oferta_sugerida({"categoria": nicho}))

    # Por oferta sugerida
    row += 2
    ws_resumo.cell(row=row, column=1, value="POR OFERTA SUGERIDA").font = Font(bold=True, size=12, color="1F2937")
    row += 1
    for col, header in enumerate(["Oferta", "Quantidade", "Leads Alta"], 1):
        ws_resumo.cell(row=row, column=col, value=header).font = resumo_header_font
        ws_resumo.cell(row=row, column=col).fill = resumo_header_fill

    ofertas = {}
    for l in leads_ordenados:
        oferta = l.get("oferta_sugerida", "")
        if oferta not in ofertas:
            ofertas[oferta] = {"total": 0, "alta": 0}
        ofertas[oferta]["total"] += 1
        if l.get("prioridade") == "Alta":
            ofertas[oferta]["alta"] += 1

    for oferta, dados in sorted(ofertas.items(), key=lambda x: x[1]["total"], reverse=True):
        row += 1
        ws_resumo.cell(row=row, column=1, value=oferta)
        ws_resumo.cell(row=row, column=2, value=dados["total"])
        ws_resumo.cell(row=row, column=3, value=dados["alta"])

    # Ajusta larguras da aba Resumo
    ws_resumo.column_dimensions["A"].width = 30
    ws_resumo.column_dimensions["B"].width = 15
    ws_resumo.column_dimensions["C"].width = 12
    ws_resumo.column_dimensions["D"].width = 12
    ws_resumo.column_dimensions["E"].width = 12
    ws_resumo.column_dimensions["F"].width = 25

    # Salva
    wb.save(caminho_saida)
    return caminho_saida


# ── Processamento Principal ───────────────────────────────────────

def processar_leads(leads):
    print(f"\n  Processando {len(leads)} leads...")

    leads_processados = []
    for lead in leads:
        score = calcular_score(lead)
        prioridade = classificar_prioridade(score)
        motivo = gerar_motivo_prioridade(lead, score)
        oferta = gerar_oferta_sugerida(lead)
        mensagem = gerar_mensagem_whatsapp(lead)
        link = gerar_link_whatsapp(lead, mensagem)

        lead["score"] = score
        lead["prioridade"] = prioridade
        lead["motivo_prioridade"] = motivo
        lead["oferta_sugerida"] = oferta
        lead["mensagem_whatsapp"] = mensagem
        lead["link_whatsapp"] = link

        leads_processados.append(lead)

    return leads_processados


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Prospecção de Leads - Gerador de Planilha Qualificada")
    parser.add_argument("--arquivo", help="Caminho para o CSV de entrada (padrão: busca automaticamente)")
    parser.add_argument("--saida", default="leads_prospeccao.xlsx", help="Nome do arquivo Excel de saída")
    args = parser.parse_args()

    # Força UTF-8 no Windows
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    print("=" * 60)
    print("  PROSPECÇÃO DE LEADS - GERADOR DE PLANILHA QUALIFICADA")
    print("=" * 60)

    arquivo = encontrar_arquivo_dados(args.arquivo)
    print(f"\n  Arquivo: {arquivo}")
    print(f"  Tamanho: {arquivo.stat().st_size / 1024:.1f} KB")

    leads = carregar_dados(arquivo)
    print(f"  Leads carregados: {len(leads)}")

    leads = processar_leads(leads)

    # Estatísticas rápidas
    alta = sum(1 for l in leads if l.get("prioridade") == "Alta")
    media = sum(1 for l in leads if l.get("prioridade") == "Média")
    baixa = sum(1 for l in leads if l.get("prioridade") == "Baixa")
    com_whatsapp = sum(1 for l in leads if l.get("link_whatsapp"))

    print(f"\n  RESULTADO:")
    print(f"  Alta prioridade:  {alta}")
    print(f"  Média prioridade: {media}")
    print(f"  Baixa prioridade: {baixa}")
    print(f"  Com link WhatsApp: {com_whatsapp}")

    # Exporta
    PROSPECCAO_DIR.mkdir(parents=True, exist_ok=True)
    caminho_saida = PROSPECCAO_DIR / args.saida
    exportar_excel(leads, caminho_saida)

    print(f"\n  ✓ Excel exportado: {caminho_saida}")
    print(f"{'=' * 60}")

    # Top 10 leads
    print(f"\n  TOP 10 LEADS (Alta Prioridade):")
    print(f"  {'-'*55}")
    top = sorted(leads, key=lambda l: l.get("score", 0), reverse=True)[:10]
    for i, lead in enumerate(top, 1):
        pri = lead.get("prioridade", "?")
        score = lead.get("score", 0)
        nome = lead.get("nome", "?")[:35]
        cidade = lead.get("cidade", "?")[:15]
        oferta = lead.get("oferta_sugerida", "?")
        print(f"  {i:2d}. [{pri:5s}] {score:3d}pts | {nome:35s} | {cidade:15s} | {oferta}")

    print()


if __name__ == "__main__":
    main()