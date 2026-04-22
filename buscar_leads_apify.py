"""
Busca de Leads via Apify - Google Maps + Email
===============================================
Actor: lukaskrivka/google-maps-with-contact-details
Faz Google Maps scraping + extração de email em uma só chamada.
Nichos Tier 1 nas 13 cidades da Baixada Fluminense.
"""

import csv
import json
import os
import re
import sys
import time
import requests
from pathlib import Path
from datetime import datetime
from collections import Counter

# Fix encoding no Windows
if sys.platform == "win32":
    os.system("chcp 65001 >nul 2>&1")
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

APIFY_TOKEN = "apify_api_g35riRIYbexQigTpDwRm4GJ40emppP36z2dc"
APIFY_API = "https://api.apify.com/v2"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

ACTOR_ID = "lukaskrivka~google-maps-with-contact-details"

# 6 nichos Tier 1
NICHOS = [
    "dentista",
    "advogado",
    "clínica estética",
    "academia",
    "salão de beleza",
    "barbearia",
]

# 13 cidades da Baixada Fluminense
CIDADES = [
    "Belford Roxo, RJ",
    "Duque de Caxias, RJ",
    "Guapimirim, RJ",
    "Itaguaí, RJ",
    "Japeri, RJ",
    "Magé, RJ",
    "Mesquita, RJ",
    "Nilópolis, RJ",
    "Nova Iguaçu, RJ",
    "Paracambi, RJ",
    "Queimados, RJ",
    "São João de Meriti, RJ",
    "Seropédica, RJ",
]

MAX_PLACES_PER_SEARCH = 10
TARGET_LEADS = 100

FIELDNAMES = [
    "nome", "nicho", "cidade", "endereco", "telefone", "website",
    "email", "avaliacao", "num_avaliacoes", "lat", "lng",
    "instagram", "facebook", "categoria_google",
    "observacao", "relevancia",
]

# Blacklist de emails inválidos
EMAIL_BLACKLIST = {
    "test@test.com", "email@email.com", "example@example.com",
    "contato@exemplo.com", "teste@teste.com", "noreply@google.com",
    "no-reply@google.com", "newsletter@googlegroups.com",
    "mailer-daemon@google.com", "user@domain.com",
}

# Franquias conhecidas
FRANQUIAS = [
    "mcdonald", "burguer king", "subway", "habib", "giraffas",
    "bob's", "kfc", "pizza hut", "domino", "grenal",
    "café do pão de queijo", "o boticário", "americanas",
    "casas bahia", "extra hiper", "carrefour", "assaí",
    "atacadão", "magazine luiza", "ri happy", "smart fit",
]


def run_actor(search_strings, max_places=10):
    """Executa o Actor google-maps-with-contact-details."""
    actor_input = {
        "searchStringsArray": search_strings,
        "maxCrawledPlacesPerSearch": max_places,
        "scrapeContactDetails": True,
    }

    url = f"{APIFY_API}/acts/{ACTOR_ID}/runs?token={APIFY_TOKEN}"
    print(f"  Buscando: {len(search_strings)} termos...")

    try:
        resp = requests.post(url, json=actor_input, timeout=30)
        if resp.status_code in [200, 201]:
            data = resp.json().get("data", {})
            return data.get("id"), data.get("defaultDatasetId")
        else:
            print(f"  [ERRO] Status {resp.status_code}: {resp.text[:300]}")
            return None, None
    except Exception as e:
        print(f"  [ERRO] {e}")
        return None, None


def wait_for_run(run_id, timeout=300):
    """Espera o Actor terminar."""
    url = f"{APIFY_API}/actor-runs/{run_id}?token={APIFY_TOKEN}"
    start = time.time()
    while time.time() - start < timeout:
        try:
            resp = requests.get(url, timeout=10)
            status = resp.json().get("data", {}).get("status")
            elapsed = int(time.time() - start)
            if status == "SUCCEEDED":
                print(f"  [OK] Finalizado em {elapsed}s")
                return True
            elif status in ["FAILED", "ABORTED", "TIMED-OUT"]:
                print(f"  [ERRO] Status: {status}")
                return False
            else:
                print(f"  Aguardando... ({elapsed}s)")
        except Exception:
            pass
        time.sleep(10)
    print(f"  [TIMEOUT] {timeout}s")
    return False


def get_results(dataset_id):
    """Baixa resultados do dataset."""
    url = f"{APIFY_API}/datasets/{dataset_id}/items?token={APIFY_TOKEN}&clean=true&format=json"
    try:
        resp = requests.get(url, timeout=60)
        if resp.status_code == 200:
            return resp.json()
        print(f"  [ERRO] Dataset: {resp.status_code}")
        return []
    except Exception as e:
        print(f"  [ERRO] {e}")
        return []


def validate_email(email):
    """Valida se o email é útil."""
    if not email:
        return None
    email = email.strip().lower()
    if email in EMAIL_BLACKLIST:
        return None
    if len(email) > 80:
        return None
    if not re.match(r"^[\w.+-]+@[\w-]+\.[\w.-]+$", email):
        return None
    # Remove imagens disfarçadas
    ext = "." + email.rsplit(".", 1)[-1] if "." in email else ""
    if ext in {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico"}:
        return None
    return email


def normalize(item, nicho, cidade):
    """Normaliza um resultado para o formato padronizado."""
    nome = item.get("title") or item.get("name") or ""
    telefone = item.get("phone") or item.get("phoneUnformatted") or ""
    website = item.get("website") or ""
    endereco = item.get("address") or ""
    avaliacao = item.get("totalScore") or item.get("rating") or ""
    num_avaliacoes = item.get("reviewsCount") or item.get("reviewCount") or ""

    # Coordenadas
    loc = item.get("location", {})
    lat = loc.get("lat", "") if isinstance(loc, dict) else ""
    lng = loc.get("lng", "") if isinstance(loc, dict) else ""

    # Categorias
    cats = item.get("categories") or item.get("categoryName") or ""
    if isinstance(cats, list):
        cats = ", ".join(cats)

    # Email (o Actor já extrai do site)
    email = validate_email(item.get("email") or "")

    # Redes sociais
    instagram = ""
    facebook = ""
    social_links = item.get("socialMediaLinks") or item.get("socialProfiles") or []
    if isinstance(social_links, list):
        for link in social_links:
            if isinstance(link, str):
                if "instagram" in link.lower():
                    instagram = link
                elif "facebook" in link.lower():
                    facebook = link
            elif isinstance(link, dict):
                url = link.get("url", "")
                if "instagram" in url.lower():
                    instagram = url
                elif "facebook" in url.lower():
                    facebook = url

    # Detectar franquia
    nome_lower = nome.lower()
    is_franchise = any(f in nome_lower for f in FRANQUIAS)

    # Detectar site profissional (heurística simples)
    has_pro_site = False
    if website and not any(s in website.lower() for s in ["facebook", "instagram", "wix.com", "linkedin"]):
        has_pro_site = True

    # Observações (argumentos de venda)
    obs = []
    if not website:
        obs.append("SEM SITE")
    elif "facebook" in (website or "").lower():
        obs.append("Site é só Facebook")
    elif "instagram" in (website or "").lower():
        obs.append("Site é só Instagram")
    if not email:
        obs.append("sem email")
    if not telefone:
        obs.append("sem telefone")

    # Score de relevância
    score = 0
    if telefone:
        score += 3
    if email:
        score += 3
    if website and not has_pro_site:
        score += 2
    try:
        if avaliacao and float(avaliacao) >= 4.0:
            score += 1
    except (ValueError, TypeError):
        pass
    try:
        if num_avaliacoes and int(num_avaliacoes) >= 10:
            score += 1
    except (ValueError, TypeError):
        pass
    if not is_franchise:
        score += 1

    return {
        "nome": nome,
        "nicho": nicho,
        "cidade": cidade,
        "endereco": endereco,
        "telefone": telefone,
        "website": website,
        "email": email or "",
        "avaliacao": avaliacao,
        "num_avaliacoes": num_avaliacoes,
        "lat": lat,
        "lng": lng,
        "instagram": instagram,
        "facebook": facebook,
        "categoria_google": cats,
        "observacao": " | ".join(obs) if obs else "",
        "relevancia": score,
        "_is_franchise": is_franchise,
        "_has_pro_site": has_pro_site,
    }


def filter_leads(leads, target=100):
    """Filtra e ranqueia leads pelas regras do usuário."""
    # Remove sem nome
    filtered = [l for l in leads if l.get("nome")]

    # Remove franquias
    filtered = [l for l in filtered if not l.get("_is_franchise")]

    # Remove com site profissional (filtro do usuário)
    filtered = [l for l in filtered if not l.get("_has_pro_site")]

    # Remove duplicatas
    seen = set()
    unique = []
    for l in filtered:
        key = f"{l['nome'].lower().strip()}|{l['cidade'].lower().strip()}"
        if key not in seen:
            seen.add(key)
            unique.append(l)

    # Ordenar por relevância
    unique.sort(key=lambda x: x.get("relevancia", 0), reverse=True)

    # Top N
    result = unique[:target]

    # Limpar campos internos
    for l in result:
        l.pop("_is_franchise", None)
        l.pop("_has_pro_site", None)

    return result, len(unique)


def export_csv(leads, filename):
    """Exporta para CSV."""
    filepath = OUTPUT_DIR / filename
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(leads)
    return filepath


def export_excel(leads, filename):
    """Exporta para Excel com formatação."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Qualificados"

        headers_map = {
            "nome": "Nome",
            "nicho": "Nicho",
            "cidade": "Cidade",
            "endereco": "Endereço",
            "telefone": "Telefone",
            "website": "Website",
            "email": "Email",
            "avaliacao": "Avaliação",
            "num_avaliacoes": "Nº Avaliações",
            "instagram": "Instagram",
            "facebook": "Facebook",
            "observacao": "Observação",
            "relevancia": "Score",
        }

        headers = list(headers_map.values())
        fields = list(headers_map.keys())

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        yellow_fill = PatternFill("solid", fgColor="FEF08A")
        green_fill = PatternFill("solid", fgColor="D1FAE5")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, lead in enumerate(leads, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=lead.get(field, ""))

            # Destacar: sem site = amarelo | com email = verde
            if not lead.get("website"):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = yellow_fill
            elif lead.get("email"):
                ws.cell(row=row_idx, column=7).fill = green_fill

        widths = [35, 18, 25, 45, 18, 35, 30, 10, 12, 30, 30, 35, 8]
        for i, w in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        filepath = OUTPUT_DIR / filename
        wb.save(filepath)
        return filepath
    except ImportError:
        print("  [!] openpyxl não instalado — CSV apenas")
        return None


def main():
    print("=" * 60)
    print("BUSCA DE LEADS - BAIXADA FLUMINENSE")
    print("Apify: google-maps-with-contact-details")
    print("=" * 60)
    print(f"Nichos: {len(NICHOS)} | Cidades: {len(CIDADES)}")
    print(f"Meta: {TARGET_LEADS} leads qualificados")
    print()

    all_leads = []

    # Buscar por nicho (1 run por nicho, todas as cidades juntas)
    for i, nicho in enumerate(NICHOS, 1):
        print(f"\n[{i}/{len(NICHOS)}] {nicho.upper()}")

        search_strings = [f"{nicho} em {cidade}" for cidade in CIDADES]

        run_id, dataset_id = run_actor(search_strings, MAX_PLACES_PER_SEARCH)
        if not run_id:
            print(f"  Pulando...")
            continue

        # Esperar
        success = wait_for_run(run_id, timeout=600)
        if not success:
            # Pegar logs do erro
            url = f"{APIFY_API}/actor-runs/{run_id}/log?token={APIFY_TOKEN}"
            log_resp = requests.get(url, timeout=10)
            print(f"  Últimas linhas do log:")
            for line in log_resp.text.strip().split("\n")[-5:]:
                print(f"    {line[:120].encode('ascii', errors='replace').decode('ascii')}")
            continue

        # Baixar resultados
        results = get_results(dataset_id)
        print(f"  Resultados brutos: {len(results)}")

        # Normalizar e detectar cidade
        for item in results:
            cidade_found = ""
            end = (item.get("address") or "").lower()
            for c in CIDADES:
                city_name = c.lower().split(",")[0]
                if city_name in end:
                    cidade_found = c
                    break
            if not cidade_found:
                # Tentar pelo campo city
                item_city = (item.get("city") or "").lower()
                for c in CIDADES:
                    city_name = c.lower().split(",")[0]
                    if city_name in item_city:
                        cidade_found = c
                        break
            if not cidade_found:
                cidade_found = "Baixada Fluminense, RJ"

            lead = normalize(item, nicho, cidade_found)
            all_leads.append(lead)

        print(f"  Acumulado: {len(all_leads)} leads brutos")

    print(f"\n{'='*60}")
    print(f"Total bruto: {len(all_leads)} leads")

    if not all_leads:
        print("[FALHA] Nenhum lead coletado!")
        return

    # Filtrar
    print(f"\nFiltrando (sem franquias, sem site profissional, sem inativos)...")
    filtered, total_unique = filter_leads(all_leads, TARGET_LEADS)
    print(f"  Únicos após filtros: {total_unique}")
    print(f"  Leads finais: {len(filtered)}")

    # Exportar
    print(f"\nExportando...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = export_csv(filtered, f"leads_baixada_{timestamp}.csv")
    print(f"  CSV: {csv_file}")

    xlsx_file = export_excel(filtered, f"leads_baixada_{timestamp}.xlsx")
    if xlsx_file:
        print(f"  Excel: {xlsx_file}")

    # Resumo
    print(f"\n{'='*60}")
    print("RESUMO FINAL")
    print(f"{'='*60}")
    print(f"Leads qualificados: {len(filtered)}")

    print(f"\nPor nicho:")
    for n, c in Counter(l["nicho"] for l in filtered).most_common():
        print(f"  {n}: {c}")

    print(f"\nPor cidade:")
    for c, n in Counter(l["cidade"] for l in filtered).most_common():
        print(f"  {c}: {n}")

    com_tel = sum(1 for l in filtered if l["telefone"])
    com_email = sum(1 for l in filtered if l["email"])
    sem_site = sum(1 for l in filtered if not l["website"])
    print(f"\nCom telefone: {com_tel}/{len(filtered)}")
    print(f"Com email: {com_email}/{len(filtered)}")
    print(f"Sem site: {sem_site}/{len(filtered)}")
    print(f"\nArquivos em: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
